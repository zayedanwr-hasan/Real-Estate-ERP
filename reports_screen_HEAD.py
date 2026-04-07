import os
import tkinter as tk
from datetime import date, datetime, timedelta
from decimal import Decimal
from tkinter import filedialog, messagebox, ttk
from typing import Any, Dict, List, Sequence, Tuple

import ttkbootstrap as tb

from combobox_helper import bind_searchable_combobox, set_combobox_values
from db_connection import get_connection, get_db_error_message


# =========================================================
# Export helpers
# =========================================================

def export_to_excel(data, filename):
    """
    Export a report payload to Excel using pandas/openpyxl.
    Also formats headers, freezes panes, and autosizes columns.
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise RuntimeError("┘à┘â╪¬╪¿╪⌐ pandas/openpyxl ╪║┘è╪▒ ┘à╪½╪¿╪¬╪⌐. ╪½╪¿┘æ╪¬┘ç╪º ╪ú┘ê┘ä╪º┘ï.") from exc

    df = pd.DataFrame(data.get("rows", []), columns=data.get("columns", []))

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        sheet_name = (data.get("title") or "Report")[:31]
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]]

    # Header style
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze first row
    ws.freeze_panes = "A2"

    # Autosize columns
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 35)

    wb.save(filename)


def export_to_pdf(data, filename, title):
    """
    Export a report payload to PDF using reportlab.
    Supports optional Arabic shaping if arabic_reshaper + bidi are installed.
    Attempts to use a Unicode font available on the system.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        raise RuntimeError("┘à┘â╪¬╪¿╪⌐ reportlab ╪║┘è╪▒ ┘à╪½╪¿╪¬╪⌐. ╪½╪¿┘æ╪¬┘ç╪º ╪ú┘ê┘ä╪º┘ï.") from exc

    def shape_text(text):
        text = "" if text is None else str(text)
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(text))
        except Exception:
            return text

    def find_unicode_font():
        candidates = [
            r"C:\Windows\Fonts\arial.ttf",
            r"C:\Windows\Fonts\segoeui.ttf",
            r"C:\Windows\Fonts\tahoma.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        ]
        for path in candidates:
            if os.path.exists(path):
                return path
        return None

    font_name = "Helvetica"
    font_path = find_unicode_font()
    if font_path:
        try:
            font_name = "AppUnicodeFont"
            pdfmetrics.registerFont(TTFont(font_name, font_path))
        except Exception:
            font_name = "Helvetica"

    doc = SimpleDocTemplate(
        filename,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="ArabicTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=16,
        leading=20,
        alignment=2,
    ))
    styles.add(ParagraphStyle(
        name="ArabicNormal",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        leading=13,
        alignment=2,
    ))
    styles.add(ParagraphStyle(
        name="ArabicHeading",
        parent=styles["Heading3"],
        fontName=font_name,
        fontSize=11,
        leading=14,
        alignment=2,
    ))

    story = []
    story.append(Paragraph(shape_text(title), styles["ArabicTitle"]))
    story.append(Spacer(1, 6))

    date_from = data.get("date_from", "")
    date_to = data.get("date_to", "")
    if date_from and date_to:
        story.append(Paragraph(shape_text(f"╪º┘ä┘ü╪¬╪▒╪⌐: {date_from} ╪Ñ┘ä┘ë {date_to}"), styles["ArabicNormal"]))
        story.append(Spacer(1, 6))

    columns = [shape_text(c) for c in data.get("columns", [])]
    rows = data.get("rows", [])

    def fmt(v):
        if v is None:
            return ""
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, (int, float, Decimal)):
            try:
                return f"{float(v):,.2f}"
            except Exception:
                return str(v)
        return shape_text(v)

    table_data = [columns] + [[fmt(v) for v in row] for row in rows]

    if not rows:
        story.append(Paragraph(shape_text("┘ä╪º ╪¬┘ê╪¼╪» ╪¿┘è╪º┘å╪º╪¬ ┘ä╪╣╪▒╪╢┘ç╪º."), styles["ArabicHeading"]))
        doc.build(story)
        return

    col_count = len(columns)
    base_width = 720 / max(col_count, 1)
    col_widths = [base_width for _ in range(col_count)]

    report_table = Table(table_data, repeatRows=1, colWidths=col_widths)
    report_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d8e0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7f6")]),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(report_table)

    summary = data.get("summary", {})
    if summary:
        story.append(Spacer(1, 8))
        summary_text = " | ".join([f"{shape_text(k)}: {fmt(v)}" for k, v in summary.items()])
        story.append(Paragraph(shape_text(summary_text), styles["ArabicHeading"]))

    doc.build(story)


# =========================================================
# Main Reports Screen
# =========================================================

class ReportsScreen:
    def __init__(self, master):
        self.master = master

        self.primary_color = "#2C3E50"
        self.sidebar_color = "#34495E"
        self.accent_color = "#1ABC9C"
        self.soft_bg = "#F4F7F6"
        self.card_bg = "#FFFFFF"
        self.soft_card_bg = "#F8FAFB"
        self.summary_bg = "#EAF7F4"
        self.text_color = "#1F2D3D"

        # Match typography used in sub_coding screens.
        self.base_font = ("Segoe UI", 11)
        self.bold_font = ("Segoe UI", 11, "bold")
        self.title_font = ("Segoe UI", 20, "bold")

        self.report_to_route = {
            "┘â╪┤┘ü ╪¡╪│╪º╪¿ ┘ê╪º╪▒╪½": ("heirs", "┘â╪┤┘ü ╪¡╪│╪º╪¿ ┘ê╪º╪▒╪½"),
            "┘à┘ä╪«╪╡ ┘ê╪º╪▒╪½": ("heirs", "┘à┘ä╪«╪╡ ┘ê╪º╪▒╪½"),
            "╪ú╪▒╪╡╪»╪⌐ ╪º┘ä┘ê╪▒╪½╪⌐": ("heirs", "╪ú╪▒╪╡╪»╪⌐ ╪º┘ä┘ê╪▒╪½╪⌐"),
            "╪ú╪▒╪╡╪»╪⌐ ╪º┘ä┘ê╪▒╪½╪⌐ ╪º┘ê ╪º┘ä┘à╪¼┘à┘ê╪╣╪⌐": ("heirs", "╪ú╪▒╪╡╪»╪⌐ ╪º┘ä┘ê╪▒╪½╪⌐"),
            "╪¬┘ê╪▓┘è╪╣ ╪º┘ä╪ú╪▒╪¿╪º╪¡": ("heirs", "╪¬┘ê╪▓┘è╪╣ ╪º┘ä╪ú╪▒╪¿╪º╪¡"),

            "┘â╪┤┘ü ╪¡╪│╪º╪¿ ╪╣┘é╪º╪▒": ("properties", "┘â╪┤┘ü ╪¡╪│╪º╪¿ ╪╣┘é╪º╪▒"),
            "╪¬┘é╪▒┘è╪▒ ╪º┘ä╪Ñ┘è╪▒╪º╪»╪º╪¬": ("properties", "╪¬┘é╪▒┘è╪▒ ╪º┘ä╪Ñ┘è╪▒╪º╪»╪º╪¬"),
            "╪¬┘é╪▒┘è╪▒ ╪º┘ä┘à╪╡╪▒┘ê┘ü╪º╪¬": ("properties", "╪¬┘é╪▒┘è╪▒ ╪º┘ä┘à╪╡╪▒┘ê┘ü╪º╪¬"),
            "╪╡╪º┘ü┘è ╪º┘ä╪▒╪¿╪¡": ("properties", "╪╡╪º┘ü┘è ╪º┘ä╪▒╪¿╪¡"),

            "╪│┘å╪»╪º╪¬ ╪º┘ä╪╡╪▒┘ü": ("vouchers", "╪│┘å╪»╪º╪¬ ╪º┘ä╪╡╪▒┘ü"),
            "╪│┘å╪»╪º╪¬ ╪º┘ä┘é╪¿╪╢": ("vouchers", "╪│┘å╪»╪º╪¬ ╪º┘ä┘é╪¿╪╢"),
            "╪¬┘é╪▒┘è╪▒ ┘è┘ê┘à┘è": ("vouchers", "╪¬┘é╪▒┘è╪▒ ┘è┘ê┘à┘è"),

            "╪»┘ü╪¬╪▒ ╪º┘ä╪ú╪│╪¬╪º╪░": ("financial", "╪»┘ü╪¬╪▒ ╪º┘ä╪ú╪│╪¬╪º╪░"),
            "┘à┘è╪▓╪º┘å ╪º┘ä┘à╪▒╪º╪¼╪╣╪⌐": ("financial", "┘à┘è╪▓╪º┘å ╪º┘ä┘à╪▒╪º╪¼╪╣╪⌐"),
            "┘é╪º╪ª┘à╪⌐ ╪º┘ä╪»╪«┘ä": ("financial", "┘é╪º╪ª┘à╪⌐ ╪º┘ä╪»╪«┘ä"),
            "╪º┘ä╪¬╪»┘ü┘é╪º╪¬ ╪º┘ä┘å┘é╪»┘è╪⌐": ("financial", "╪º┘ä╪¬╪»┘ü┘é╪º╪¬ ╪º┘ä┘å┘é╪»┘è╪⌐"),

            "╪¡╪│╪¿ ╪º┘ä╪¡╪│╪º╪¿": ("advanced", "╪¡╪│╪¿ ╪º┘ä╪¡╪│╪º╪¿"),
            "╪¡╪│╪¿ ╪º┘ä┘ê╪º╪▒╪½": ("advanced", "╪¡╪│╪¿ ╪º┘ä┘ê╪º╪▒╪½"),
            "╪¡╪│╪¿ ╪º┘ä╪╣┘é╪º╪▒": ("advanced", "╪¡╪│╪¿ ╪º┘ä╪╣┘é╪º╪▒"),
            "┘à┘é╪º╪▒┘å╪⌐ ┘ü╪¬╪▒╪º╪¬": ("advanced", "┘à┘é╪º╪▒┘å╪⌐ ┘ü╪¬╪▒╪º╪¬"),
        }

        self.tab_order = ["heirs", "properties", "vouchers", "financial", "advanced"]
        self.tabs = {}

        self._setup_styles()
        self._build_root()
        self._build_header()
        self._build_notebook()
        self._load_filter_sources()

    # =====================================================
    # UI
    # =====================================================

    def _setup_styles(self):
        style = tb.Style()
        try:
            style.theme_use(style.theme_use())
        except Exception:
            pass

        try:
            style.configure("App.Reports.TNotebook", background=self.soft_bg, borderwidth=0)
            style.configure(
                "App.Reports.TNotebook.Tab",
                font=self.bold_font,
                padding=(14, 8),
                background=self.sidebar_color,
                foreground="#ECF0F1",
            )
            style.map(
                "App.Reports.TNotebook.Tab",
                background=[("selected", self.accent_color), ("active", self.accent_color)],
                foreground=[("selected", "white"), ("active", "white")],
            )

            style.configure("App.Reports.Treeview", rowheight=30, font=self.base_font)
            style.configure("App.Reports.Treeview.Heading", font=self.bold_font)
            style.map("App.Reports.Treeview", background=[("selected", self.accent_color)], foreground=[("selected", "white")])
        except Exception:
            pass

    def _build_root(self):
        self.frame = tk.Frame(self.master, bg=self.soft_bg)
        self.frame.pack(fill="both", expand=True)

        self.main_card = tk.Frame(
            self.frame,
            bg=self.card_bg,
            highlightthickness=1,
            highlightbackground="#D6DEE5",
        )
        self.main_card.pack(fill="both", expand=True, padx=14, pady=14)

    def _build_header(self):
        header = tk.Frame(self.main_card, bg=self.primary_color, height=58)
        header.pack(fill="x")
        header.pack_propagate(False)

        title = tk.Label(
            header,
            text="╪┤╪º╪┤╪⌐ ╪º┘ä╪¬┘é╪º╪▒┘è╪▒",
            bg=self.primary_color,
            fg="white",
            font=self.title_font,
        )
        title.pack(side="right", padx=20, pady=10)

        subtitle = tk.Label(
            header,
            text="Reports Dashboard",
            bg=self.primary_color,
            fg="#DDE6EE",
            font=self.base_font,
        )
        subtitle.pack(side="right", padx=10, pady=14)

    def _build_notebook(self):
        container = tk.Frame(self.main_card, bg=self.soft_bg)
        container.pack(fill="both", expand=True, padx=12, pady=12)

        self.notebook = ttk.Notebook(container, style="App.Reports.TNotebook")
        self.notebook.pack(fill="both", expand=True)

        self._create_report_tab(
            "heirs",
            "╪¬┘é╪º╪▒┘è╪▒ ╪º┘ä┘ê╪▒╪½╪⌐",
            {
                "┘â╪┤┘ü ╪¡╪│╪º╪¿ ┘ê╪º╪▒╪½": self.get_heir_statement,
                "┘à┘ä╪«╪╡ ┘ê╪º╪▒╪½": self.get_heir_summary,
                "╪ú╪▒╪╡╪»╪⌐ ╪º┘ä┘ê╪▒╪½╪⌐": self.get_heir_balances,
                "╪¬┘ê╪▓┘è╪╣ ╪º┘ä╪ú╪▒╪¿╪º╪¡": self.get_profit_distribution,
            },
        )

        self._create_report_tab(
            "properties",
            "╪¬┘é╪º╪▒┘è╪▒ ╪º┘ä╪╣┘é╪º╪▒╪º╪¬",
            {
                "┘â╪┤┘ü ╪¡╪│╪º╪¿ ╪╣┘é╪º╪▒": self.get_property_statement,
                "╪¬┘é╪▒┘è╪▒ ╪º┘ä╪Ñ┘è╪▒╪º╪»╪º╪¬": self.get_property_income,
                "╪¬┘é╪▒┘è╪▒ ╪º┘ä┘à╪╡╪▒┘ê┘ü╪º╪¬": self.get_property_expense,
                "╪╡╪º┘ü┘è ╪º┘ä╪▒╪¿╪¡": self.get_property_profit,
            },
        )

        self._create_report_tab(
            "vouchers",
            "╪¬┘é╪º╪▒┘è╪▒ ╪º┘ä╪│┘å╪»╪º╪¬",
            {
                "╪│┘å╪»╪º╪¬ ╪º┘ä╪╡╪▒┘ü": self.get_payment_vouchers,
                "╪│┘å╪»╪º╪¬ ╪º┘ä┘é╪¿╪╢": self.get_receipt_vouchers,
                "╪¬┘é╪▒┘è╪▒ ┘è┘ê┘à┘è": self.get_daily_vouchers,
            },
        )

        self._create_report_tab(
            "financial",
            "╪º┘ä╪¬┘é╪º╪▒┘è╪▒ ╪º┘ä┘à╪º┘ä┘è╪⌐",
            {
                "╪»┘ü╪¬╪▒ ╪º┘ä╪ú╪│╪¬╪º╪░": self.get_general_ledger,
                "┘à┘è╪▓╪º┘å ╪º┘ä┘à╪▒╪º╪¼╪╣╪⌐": self.get_trial_balance,
                "┘é╪º╪ª┘à╪⌐ ╪º┘ä╪»╪«┘ä": self.get_income_statement,
                "╪º┘ä╪¬╪»┘ü┘é╪º╪¬ ╪º┘ä┘å┘é╪»┘è╪⌐": self.get_cash_flow,
            },
        )

        self._create_report_tab(
            "advanced",
            "╪¬┘é╪º╪▒┘è╪▒ ┘à╪¬┘é╪»┘à╪⌐",
            {
                "╪¡╪│╪¿ ╪º┘ä╪¡╪│╪º╪¿": self.filter_by_account,
                "╪¡╪│╪¿ ╪º┘ä┘ê╪º╪▒╪½": self.filter_by_heir,
                "╪¡╪│╪¿ ╪º┘ä╪╣┘é╪º╪▒": self.filter_by_property,
                "┘à┘é╪º╪▒┘å╪⌐ ┘ü╪¬╪▒╪º╪¬": self.compare_periods,
            },
        )

    def _create_report_tab(self, tab_key, tab_title, reports_map):
        tab_frame = tk.Frame(self.notebook, bg=self.soft_bg)
        self.notebook.add(tab_frame, text=tab_title)

        state = {
            "reports": reports_map,
            "report_var": tk.StringVar(value=next(iter(reports_map.keys()))),
            "date_from_var": tk.StringVar(value=self._default_date_from()),
            "date_to_var": tk.StringVar(value=date.today().strftime("%Y-%m-%d")),
            "heir_var": tk.StringVar(),
            "property_var": tk.StringVar(),
            "account_var": tk.StringVar(),
            "last_data": None,
            "tree": None,
            "summary_labels": {},
        }

        filter_card = self._card(tab_frame)
        filter_card.pack(fill="x", padx=10, pady=(10, 8))

        self._build_filter_widgets(filter_card, state)
        self._build_action_buttons(filter_card, tab_key)

        table_card = self._card(tab_frame)
        table_card.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        state["tree"] = self._build_results_tree(table_card)

        summary_card = self._summary_card(tab_frame)
        summary_card.pack(fill="x", padx=10, pady=(0, 10))
        state["summary_labels"] = self._build_summary_panel(summary_card)

        self.tabs[tab_key] = state

    def _card(self, parent):
        return tk.Frame(
            parent,
            bg=self.soft_card_bg,
            highlightthickness=1,
            highlightbackground="#D8E1E8",
        )

    def _summary_card(self, parent):
        return tk.Frame(
            parent,
            bg=self.summary_bg,
            highlightthickness=1,
            highlightbackground="#B8DED6",
        )

    def _field_label(self, parent, text):
        return tk.Label(
            parent,
            text=text,
            bg=self.soft_card_bg,
            fg=self.primary_color,
            font=("Segoe UI", 12, "bold"),
        )

    def _build_filter_widgets(self, parent, state):
        for c in range(5):
            parent.grid_columnconfigure(c, weight=1)

        # Row 0
        self._field_label(parent, "┘à┘å ╪¬╪º╪▒┘è╪«").grid(row=0, column=0, padx=6, pady=(10, 4), sticky="e")
        self._field_label(parent, "╪Ñ┘ä┘ë ╪¬╪º╪▒┘è╪«").grid(row=0, column=1, padx=6, pady=(10, 4), sticky="e")
        self._field_label(parent, "╪º┘ä┘ê╪º╪▒╪½").grid(row=0, column=2, padx=6, pady=(10, 4), sticky="e")
        self._field_label(parent, "╪º┘ä╪╣┘é╪º╪▒").grid(row=0, column=3, padx=6, pady=(10, 4), sticky="e")
        self._field_label(parent, "╪º┘ä╪¡╪│╪º╪¿").grid(row=0, column=4, padx=6, pady=(10, 4), sticky="e")

        # Row 1
        date_from_entry = ttk.Entry(parent, textvariable=state["date_from_var"], justify="center", font=self.base_font)
        date_to_entry = ttk.Entry(parent, textvariable=state["date_to_var"], justify="center", font=self.base_font)
        heir_combo = ttk.Combobox(parent, textvariable=state["heir_var"], justify="right", font=self.base_font)
        property_combo = ttk.Combobox(parent, textvariable=state["property_var"], justify="right", font=self.base_font)
        account_combo = ttk.Combobox(parent, textvariable=state["account_var"], justify="right", font=self.base_font)

        date_from_entry.grid(row=1, column=0, padx=6, pady=(0, 8), sticky="ew")
        date_to_entry.grid(row=1, column=1, padx=6, pady=(0, 8), sticky="ew")
        heir_combo.grid(row=1, column=2, padx=6, pady=(0, 8), sticky="ew")
        property_combo.grid(row=1, column=3, padx=6, pady=(0, 8), sticky="ew")
        account_combo.grid(row=1, column=4, padx=6, pady=(0, 8), sticky="ew")

        bind_searchable_combobox(heir_combo)
        bind_searchable_combobox(property_combo)
        bind_searchable_combobox(account_combo)

        state["heir_combo"] = heir_combo
        state["property_combo"] = property_combo
        state["account_combo"] = account_combo

        # Row 2
        self._field_label(parent, "┘å┘ê╪╣ ╪º┘ä╪¬┘é╪▒┘è╪▒").grid(row=2, column=0, padx=6, pady=(6, 4), sticky="e")
        report_combo = ttk.Combobox(
            parent,
            textvariable=state["report_var"],
            state="readonly",
            justify="right",
            font=self.bold_font,
        )
        set_combobox_values(report_combo, list(state["reports"].keys()))
        report_combo.grid(row=3, column=0, columnspan=3, padx=6, pady=(0, 8), sticky="ew")

        state["report_combo"] = report_combo

    def _build_action_buttons(self, parent, tab_key):
        action_frame = tk.Frame(parent, bg=self.soft_card_bg)
        action_frame.grid(row=3, column=3, columnspan=2, padx=6, pady=(0, 8), sticky="w")

        tb.Button(
            action_frame,
            text="╪╣╪▒╪╢ ╪º┘ä╪¬┘é╪▒┘è╪▒",
            bootstyle="primary",
            command=lambda: self._generate_report(tab_key),
        ).pack(side="left", padx=4)

        tb.Button(
            action_frame,
            text="╪¬╪╡╪»┘è╪▒ PDF",
            bootstyle="warning",
            command=lambda: self._on_export_pdf(tab_key),
        ).pack(side="left", padx=4)

        tb.Button(
            action_frame,
            text="╪¬╪╡╪»┘è╪▒ Excel",
            bootstyle="success",
            command=lambda: self._on_export_excel(tab_key),
        ).pack(side="left", padx=4)

    def _build_results_tree(self, parent):
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        tree = ttk.Treeview(parent, show="headings", style="App.Reports.Treeview")
        y_scroll = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)

        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        tree.tag_configure("even", background="#F8FAFB")
        tree.tag_configure("odd", background="#FFFFFF")

        return tree

    def _build_summary_panel(self, parent):
        labels = {}
        for c in range(3):
            parent.grid_columnconfigure(c, weight=1)

        summary_items = [
            ("total_debit", "╪Ñ╪¼┘à╪º┘ä┘è ╪º┘ä┘à╪»┘è┘å"),
            ("total_credit", "╪Ñ╪¼┘à╪º┘ä┘è ╪º┘ä╪»╪º╪ª┘å"),
            ("balance", "╪º┘ä╪▒╪╡┘è╪»"),
        ]

        for idx, (key, title) in enumerate(summary_items):
            item = tk.Frame(
                parent,
                bg=self.summary_bg,
                highlightthickness=1,
                highlightbackground="#B8DED6",
                padx=12,
                pady=10,
            )
            item.grid(row=0, column=idx, padx=6, pady=8, sticky="ew")

            tk.Label(
                item,
                text=title,
                bg=self.summary_bg,
                fg=self.primary_color,
                font=("Segoe UI", 12, "bold"),
            ).pack(anchor="e")

            labels[key] = tk.Label(
                item,
                text="0.00",
                bg=self.summary_bg,
                fg="#0E6251",
                font=("Segoe UI", 20, "bold"),
            )
            labels[key].pack(anchor="e")

        return labels

    # =====================================================
    # Data sources
    # =====================================================

    def _default_date_from(self):
        today = date.today()
        return today.replace(day=1).strftime("%Y-%m-%d")

    def _load_filter_sources(self):
        vendors = self._fetch_lookup(
            "SELECT id, vendor_name FROM finance.vendors ORDER BY vendor_name",
            "╪º┘ä┘à┘ê╪▒╪»┘è┘å",
        )
        properties = self._fetch_lookup(
            "SELECT id, property_name FROM finance.properties ORDER BY property_name",
            "╪º┘ä╪╣┘é╪º╪▒╪º╪¬",
        )
        accounts = self._fetch_lookup(
            "SELECT account_code, account_name FROM finance.accounts ORDER BY account_code",
            "╪º┘ä╪¡╪│╪º╪¿╪º╪¬",
        )

        vendor_values = [f"{row[0]} - {row[1]}" for row in vendors]
        property_values = [f"{row[0]} - {row[1]}" for row in properties]
        account_values = [f"{row[0]} - {row[1]}" for row in accounts]

        for state in self.tabs.values():
            set_combobox_values(state["heir_combo"], vendor_values)
            set_combobox_values(state["property_combo"], property_values)
            set_combobox_values(state["account_combo"], account_values)

    def _fetch_lookup(self, query, source_name) -> List[Tuple[Any, ...]]:
        conn = get_connection()
        if not conn:
            return []
        try:
            with conn:
                with conn.cursor() as cur:
                    cur.execute(query)
                    return list(cur.fetchall() or [])
        except Exception as exc:
            messagebox.showerror("╪«╪╖╪ú", get_db_error_message(exc, f"╪¬╪╣╪░╪▒ ╪¬╪¡┘à┘è┘ä ╪¿┘è╪º┘å╪º╪¬ {source_name}"))
            return []
        finally:
            conn.close()

    # =====================================================
    # Helpers
    # =====================================================

    def _tab_key_from_state(self, state):
        for key, value in self.tabs.items():
            if value is state:
                return key
        return "heirs"

    def _extract_prefix(self, value):
        text = (value or "").strip()
        if not text:
            return ""
        return text.split(" - ")[0].strip() if " - " in text else text

    def _is_int_like(self, value):
        return isinstance(value, int) and not isinstance(value, bool)

    def _format_cell(self, column_name, value):
        if value is None:
            return ""

        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")

        col = (column_name or "").lower()

        if self._is_int_like(value):
            if col in {"id", "voucher_id", "vouchers_count"} or col.endswith("_id") or col.endswith("_count"):
                return f"{int(value):,}"
            return f"{int(value):,}"

        if isinstance(value, (float, Decimal)):
            if col in {
                "debit", "credit", "balance", "income", "expense",
                "cash_in", "cash_out", "total_profit", "distributed_share",
                "previous_period", "current_period", "difference",
                "net_profit", "profit"
            }:
                return f"{float(value):,.2f}"
            return f"{float(value):,.2f}"

        return str(value)

    def _numeric_value(self, value):
        if value is None:
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    def _extract_optional_int(self, value, field_name):
        token = self._extract_prefix(value)
        if not token:
            return ""
        if token.isdigit():
            return int(token)
        raise ValueError(f"┘è╪▒╪¼┘ë ╪º╪«╪¬┘è╪º╪▒ {field_name} ┘à┘å ╪º┘ä┘é╪º╪ª┘à╪⌐")

    def _validate_filters(self, tab_key):
        state = self.tabs[tab_key]
        date_from = state["date_from_var"].get().strip()
        date_to = state["date_to_var"].get().strip()

        try:
            from_dt = datetime.strptime(date_from, "%Y-%m-%d").date()
            to_dt = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            messagebox.showwarning("╪¬┘å╪¿┘è┘ç", "╪╡┘è╪║╪⌐ ╪º┘ä╪¬╪º╪▒┘è╪« ┘è╪¼╪¿ ╪ú┘å ╪¬┘â┘ê┘å YYYY-MM-DD")
            return None

        if from_dt > to_dt:
            messagebox.showwarning("╪¬┘å╪¿┘è┘ç", "╪¬╪º╪▒┘è╪« ╪º┘ä╪¿╪»╪º┘è╪⌐ ┘è╪¼╪¿ ╪ú┘å ┘è┘â┘ê┘å ┘é╪¿┘ä ╪¬╪º╪▒┘è╪« ╪º┘ä┘å┘ç╪º┘è╪⌐")
            return None

        try:
            vendor_id = self._extract_optional_int(state["heir_var"].get(), "╪º┘ä┘ê╪º╪▒╪½")
            property_id = self._extract_optional_int(state["property_var"].get(), "╪º┘ä╪╣┘é╪º╪▒")
        except ValueError as exc:
            messagebox.showwarning("╪¬┘å╪¿┘è┘ç", str(exc))
            return None

        return {
            "date_from": date_from,
            "date_to": date_to,
            "vendor_id": vendor_id,
            "property_id": property_id,
            "account_code": self._extract_prefix(state["account_var"].get()),
        }

    def _ledger_filter_clause(self, filters, date_field="v.v_date", property_field="l.property_id"):
        conditions = [f"{date_field} BETWEEN %s AND %s"]
        params = [filters["date_from"], filters["date_to"]]

        if filters.get("vendor_id"):
            conditions.append("l.vendor_id = %s")
            params.append(filters["vendor_id"])
        if filters.get("property_id"):
            conditions.append(f"{property_field} = %s")
            params.append(filters["property_id"])
        if filters.get("account_code"):
            conditions.append("l.account_code = %s")
            params.append(filters["account_code"])

        return " AND ".join(conditions), params

    def _run_query(self, query: str, params: Sequence[Any]) -> Tuple[List[str], List[Tuple[Any, ...]]]:
        conn = get_connection()
        if not conn:
            raise RuntimeError("╪¬╪╣╪░╪▒ ╪º┘ä╪º╪¬╪╡╪º┘ä ╪¿┘é╪º╪╣╪»╪⌐ ╪º┘ä╪¿┘è╪º┘å╪º╪¬")

        try:
            with conn:
                with conn.cursor() as cur:
                    cur.execute(query, params)
                    rows = list(cur.fetchall() or [])
                    columns = [desc[0] for desc in cur.description] if cur.description else []
            return columns, rows
        except Exception as exc:
            raise RuntimeError(get_db_error_message(exc, "╪«╪╖╪ú ┘ü┘è ┘é╪º╪╣╪»╪⌐ ╪º┘ä╪¿┘è╪º┘å╪º╪¬")) from exc
        finally:
            conn.close()

    def _compute_summary(self, columns, rows):
        if not rows or not columns:
            return {"total_debit": 0.0, "total_credit": 0.0, "balance": 0.0}

        idx = {name.lower(): pos for pos, name in enumerate(columns)}
        total_debit = 0.0
        total_credit = 0.0

        for row in rows:
            for col_name, pos in idx.items():
                val = self._numeric_value(row[pos])
                if col_name in {"debit", "expense", "cash_out"}:
                    total_debit += val
                elif col_name in {"credit", "income", "cash_in"}:
                    total_credit += val
                elif col_name == "profit":
                    total_credit += val
                elif col_name == "distributed_share":
                    total_credit += val

        return {
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balance": total_debit - total_credit,
        }

    def _update_tree(self, state, columns, rows):
        tree = state["tree"]
        tree.delete(*tree.get_children())

        tree["columns"] = columns

        for col in columns:
            tree.heading(col, text=col, anchor="e")
            tree.column(col, anchor="e", width=150, minwidth=110, stretch=True)

        for idx, row in enumerate(rows):
            formatted = [self._format_cell(columns[i], v) for i, v in enumerate(row)]
            tag = "even" if idx % 2 == 0 else "odd"
            tree.insert("", "end", values=formatted, tags=(tag,))

    def _update_summary(self, state, summary):
        state["summary_labels"]["total_debit"].configure(text=f"{summary.get('total_debit', 0.0):,.2f}")
        state["summary_labels"]["total_credit"].configure(text=f"{summary.get('total_credit', 0.0):,.2f}")
        state["summary_labels"]["balance"].configure(text=f"{summary.get('balance', 0.0):,.2f}")

    def _run_report_query(self, title: str, query: str, params: Sequence[Any], filters: Dict[str, Any]):
        columns, rows = self._run_query(query, params)
        summary = self._compute_summary(columns, rows)
        return {
            "title": title,
            "columns": columns,
            "rows": rows,
            "summary": summary,
            "date_from": filters["date_from"],
            "date_to": filters["date_to"],
        }

    def _normalize_payload(self, payload, filters, report_name):
        columns = payload.get("columns") or []
        rows = payload.get("rows") or []
        summary = payload.get("summary") or self._compute_summary(columns, rows)

        return {
            "title": payload.get("title") or report_name,
            "columns": columns,
            "rows": rows,
            "summary": summary,
            "date_from": payload.get("date_from") or filters.get("date_from", ""),
            "date_to": payload.get("date_to") or filters.get("date_to", ""),
        }

    def _generate_report(self, tab_key):
        filters = self._validate_filters(tab_key)
        if not filters:
            return

        state = self.tabs[tab_key]
        report_name = state["report_var"].get()
        handler = state["reports"].get(report_name)

        if not handler:
            messagebox.showwarning("╪¬┘å╪¿┘è┘ç", "┘ä┘à ┘è╪¬┘à ╪º┘ä╪╣╪½┘ê╪▒ ╪╣┘ä┘ë ┘å┘ê╪╣ ╪º┘ä╪¬┘é╪▒┘è╪▒")
            return

        try:
            payload = handler(filters)
            payload = self._normalize_payload(payload, filters, report_name)
        except ValueError as exc:
            messagebox.showwarning("╪¬┘å╪¿┘è┘ç", str(exc))
            return
        except Exception as exc:
            messagebox.showerror("╪«╪╖╪ú", f"╪¬╪╣╪░╪▒ ╪Ñ┘å╪┤╪º╪í ╪º┘ä╪¬┘é╪▒┘è╪▒: {exc}")
            return

        if not payload.get("rows"):
            self._update_tree(state, payload.get("columns", []), [])
            self._update_summary(state, payload.get("summary", {}))
            state["last_data"] = payload
            messagebox.showinfo("┘å╪¬┘è╪¼╪⌐", "┘ä╪º ╪¬┘ê╪¼╪» ╪¿┘è╪º┘å╪º╪¬ ┘à╪╖╪º╪¿┘é╪⌐ ┘ä┘ä┘ü┘ä╪º╪¬╪▒ ╪º┘ä┘à╪¡╪»╪»╪⌐")
            return

        self._update_tree(state, payload.get("columns", []), payload.get("rows", []))
        self._update_summary(state, payload.get("summary", {}))
        state["last_data"] = payload

    def _export_report(self, tab_key, extension, filetypes, exporter, success_message):
        state = self.tabs[tab_key]
        data = state.get("last_data")
        if not data or not data.get("rows"):
            messagebox.showwarning("╪¬┘å╪¿┘è┘ç", "╪º╪╣╪▒╪╢ ╪º┘ä╪¬┘é╪▒┘è╪▒ ╪ú┘ê┘ä╪º┘ï ┘é╪¿┘ä ╪º┘ä╪¬╪╡╪»┘è╪▒")
            return

        file_name = filedialog.asksaveasfilename(
            defaultextension=extension,
            filetypes=filetypes,
        )
        if not file_name:
            return

        try:
            exporter(data, file_name)
            messagebox.showinfo("┘å╪¼╪º╪¡", success_message)
        except Exception as exc:
            messagebox.showerror("╪«╪╖╪ú", str(exc))

    def _on_export_excel(self, tab_key):
        self._export_report(
            tab_key,
            ".xlsx",
            [("Excel Files", "*.xlsx")],
            lambda data, file_name: export_to_excel(data, file_name),
            "╪¬┘à ╪¬╪╡╪»┘è╪▒ ╪º┘ä╪¬┘é╪▒┘è╪▒ ╪Ñ┘ä┘ë Excel ╪¿┘å╪¼╪º╪¡",
        )

    def _on_export_pdf(self, tab_key):
        self._export_report(
            tab_key,
            ".pdf",
            [("PDF Files", "*.pdf")],
            lambda data, file_name: export_to_pdf(data, file_name, data.get("title", "╪¬┘é╪▒┘è╪▒")),
            "╪¬┘à ╪¬╪╡╪»┘è╪▒ ╪º┘ä╪¬┘é╪▒┘è╪▒ ╪Ñ┘ä┘ë PDF ╪¿┘å╪¼╪º╪¡",
        )

    # =====================================================
    # Report SQLs
    # =====================================================

    # ---------- Heirs ----------
    def get_heir_statement(self, filters):
        where_sql, params = self._ledger_filter_clause(filters, property_field="vd.property_id")
        query = f"""
            SELECT
                v.id AS voucher_id,
                v.v_date,
                COALESCE(vd.vendor_name, '-') AS vendor_name,
                COALESCE(p.property_name, '-') AS property_name,
                COALESCE(a.account_name, '-') AS account_name,
                COALESCE(l.debit, 0) AS debit,
                COALESCE(l.credit, 0) AS credit,
                COALESCE(l.debit, 0) - COALESCE(l.credit, 0) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.vendors vd ON vd.id = l.vendor_id
            LEFT JOIN finance.properties p ON p.id = vd.property_id
            LEFT JOIN finance.accounts a ON a.account_code = l.account_code
            WHERE {where_sql}
            ORDER BY v.v_date, v.id
        """
        return self._run_report_query("┘â╪┤┘ü ╪¡╪│╪º╪¿ ┘ê╪º╪▒╪½", query, params, filters)

    def get_heir_summary(self, filters):
        where_sql, params = self._ledger_filter_clause(filters, property_field="vd.property_id")
        query = f"""
            SELECT
                COALESCE(vd.vendor_name, '╪║┘è╪▒ ┘à╪¡╪»╪»') AS vendor_name,
                SUM(COALESCE(l.debit, 0)) AS debit,
                SUM(COALESCE(l.credit, 0)) AS credit,
                SUM(COALESCE(l.debit, 0) - COALESCE(l.credit, 0)) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.vendors vd ON vd.id = l.vendor_id
            WHERE {where_sql}
            GROUP BY vd.vendor_name
            ORDER BY vendor_name
        """
        return self._run_report_query("┘à┘ä╪«╪╡ ┘ê╪º╪▒╪½", query, params, filters)

    def get_heir_balances(self, filters):
        where_sql, params = self._ledger_filter_clause(filters, property_field="vd.property_id")
        query = f"""
            SELECT
                COALESCE(vd.vendor_name, '╪║┘è╪▒ ┘à╪¡╪»╪»') AS vendor_name,
                SUM(COALESCE(l.debit, 0)) AS debit,
                SUM(COALESCE(l.credit, 0)) AS credit,
                SUM(COALESCE(l.debit, 0) - COALESCE(l.credit, 0)) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.vendors vd ON vd.id = l.vendor_id
            WHERE {where_sql}
            GROUP BY vd.vendor_name
            ORDER BY balance DESC
        """
        return self._run_report_query("╪ú╪▒╪╡╪»╪⌐ ╪º┘ä┘ê╪▒╪½╪⌐", query, params, filters)

    def get_profit_distribution(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            WITH net_profit AS (
                SELECT COALESCE(SUM(COALESCE(l.credit, 0) - COALESCE(l.debit, 0)), 0) AS profit
                FROM finance.ledger l
                JOIN finance.vouchers v ON v.id = l.voucher_id
                WHERE {where_sql}
            ),
            vendors_count AS (
                SELECT NULLIF(COUNT(*), 0) AS total_vendors
                FROM finance.vendors
            )
            SELECT
                COALESCE(vd.vendor_name, '╪║┘è╪▒ ┘à╪¡╪»╪»') AS vendor_name,
                COALESCE(np.profit, 0) AS total_profit,
                COALESCE(np.profit / vc.total_vendors, 0) AS distributed_share,
                0::numeric AS debit,
                COALESCE(np.profit / vc.total_vendors, 0) AS credit
            FROM finance.vendors vd
            CROSS JOIN net_profit np
            CROSS JOIN vendors_count vc
            ORDER BY vd.vendor_name
        """
        return self._run_report_query("╪¬┘ê╪▓┘è╪╣ ╪º┘ä╪ú╪▒╪¿╪º╪¡", query, params, filters)

    # ---------- Properties ----------
    def get_property_statement(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                v.id AS voucher_id,
                v.v_date,
                COALESCE(p.property_name, '-') AS property_name,
                COALESCE(a.account_name, '-') AS account_name,
                COALESCE(l.debit, 0) AS debit,
                COALESCE(l.credit, 0) AS credit,
                COALESCE(l.credit, 0) - COALESCE(l.debit, 0) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.properties p ON p.id = l.property_id
            LEFT JOIN finance.accounts a ON a.account_code = l.account_code
            WHERE {where_sql}
            ORDER BY v.v_date, v.id
        """
        return self._run_report_query("┘â╪┤┘ü ╪¡╪│╪º╪¿ ╪╣┘é╪º╪▒", query, params, filters)

    def get_property_income(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                COALESCE(p.property_name, '╪║┘è╪▒ ┘à╪¡╪»╪»') AS property_name,
                SUM(COALESCE(l.credit, 0)) AS income,
                0::numeric AS expense,
                SUM(COALESCE(l.credit, 0)) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.properties p ON p.id = l.property_id
            WHERE {where_sql}
              AND v.v_type = '┘é╪¿╪╢'
            GROUP BY p.property_name
            ORDER BY property_name
        """
        return self._run_report_query("╪¬┘é╪▒┘è╪▒ ╪º┘ä╪Ñ┘è╪▒╪º╪»╪º╪¬", query, params, filters)

    def get_property_expense(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                COALESCE(p.property_name, '╪║┘è╪▒ ┘à╪¡╪»╪»') AS property_name,
                SUM(COALESCE(l.debit, 0)) AS expense,
                0::numeric AS income,
                SUM(COALESCE(l.debit, 0)) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.properties p ON p.id = l.property_id
            WHERE {where_sql}
              AND v.v_type = '╪╡╪▒┘ü'
            GROUP BY p.property_name
            ORDER BY property_name
        """
        return self._run_report_query("╪¬┘é╪▒┘è╪▒ ╪º┘ä┘à╪╡╪▒┘ê┘ü╪º╪¬", query, params, filters)

    def get_property_profit(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                COALESCE(p.property_name, '╪║┘è╪▒ ┘à╪¡╪»╪»') AS property_name,
                SUM(COALESCE(l.credit, 0)) AS income,
                SUM(COALESCE(l.debit, 0)) AS expense,
                SUM(COALESCE(l.credit, 0) - COALESCE(l.debit, 0)) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.properties p ON p.id = l.property_id
            WHERE {where_sql}
            GROUP BY p.property_name
            ORDER BY balance DESC
        """
        return self._run_report_query("╪╡╪º┘ü┘è ╪º┘ä╪▒╪¿╪¡", query, params, filters)

    # ---------- Vouchers ----------
    def get_payment_vouchers(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                v.id AS voucher_id,
                v.v_date,
                COALESCE(v.description, '-') AS description,
                COALESCE(p.property_name, '-') AS property_name,
                SUM(COALESCE(l.debit, 0)) AS debit,
                0::numeric AS credit
            FROM finance.vouchers v
            JOIN finance.ledger l ON l.voucher_id = v.id
            LEFT JOIN finance.properties p ON p.id = l.property_id
            WHERE {where_sql}
              AND v.v_type = '╪╡╪▒┘ü'
            GROUP BY v.id, v.v_date, v.description, p.property_name
            ORDER BY v.v_date, v.id
        """
        return self._run_report_query("╪│┘å╪»╪º╪¬ ╪º┘ä╪╡╪▒┘ü", query, params, filters)

    def get_receipt_vouchers(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                v.id AS voucher_id,
                v.v_date,
                COALESCE(v.description, '-') AS description,
                COALESCE(p.property_name, '-') AS property_name,
                0::numeric AS debit,
                SUM(COALESCE(l.credit, 0)) AS credit
            FROM finance.vouchers v
            JOIN finance.ledger l ON l.voucher_id = v.id
            LEFT JOIN finance.properties p ON p.id = l.property_id
            WHERE {where_sql}
              AND v.v_type = '┘é╪¿╪╢'
            GROUP BY v.id, v.v_date, v.description, p.property_name
            ORDER BY v.v_date, v.id
        """
        return self._run_report_query("╪│┘å╪»╪º╪¬ ╪º┘ä┘é╪¿╪╢", query, params, filters)

    def get_daily_vouchers(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                v.v_date,
                COUNT(DISTINCT v.id) AS vouchers_count,
                SUM(COALESCE(l.debit, 0)) AS debit,
                SUM(COALESCE(l.credit, 0)) AS credit,
                SUM(COALESCE(l.credit, 0) - COALESCE(l.debit, 0)) AS balance
            FROM finance.vouchers v
            JOIN finance.ledger l ON l.voucher_id = v.id
            WHERE {where_sql}
            GROUP BY v.v_date
            ORDER BY v.v_date
        """
        return self._run_report_query("╪¬┘é╪▒┘è╪▒ ┘è┘ê┘à┘è", query, params, filters)

    # ---------- Financial ----------
    def get_general_ledger(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                v.v_date,
                v.id AS voucher_id,
                COALESCE(a.account_code, '-') AS account_code,
                COALESCE(a.account_name, '-') AS account_name,
                COALESCE(l.debit, 0) AS debit,
                COALESCE(l.credit, 0) AS credit,
                COALESCE(l.debit, 0) - COALESCE(l.credit, 0) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.accounts a ON a.account_code = l.account_code
            WHERE {where_sql}
            ORDER BY v.v_date, v.id, a.account_code
        """
        return self._run_report_query("╪»┘ü╪¬╪▒ ╪º┘ä╪ú╪│╪¬╪º╪░", query, params, filters)

    def get_trial_balance(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                COALESCE(a.account_code, '-') AS account_code,
                COALESCE(a.account_name, '-') AS account_name,
                SUM(COALESCE(l.debit, 0)) AS debit,
                SUM(COALESCE(l.credit, 0)) AS credit,
                SUM(COALESCE(l.debit, 0) - COALESCE(l.credit, 0)) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.accounts a ON a.account_code = l.account_code
            WHERE {where_sql}
            GROUP BY a.account_code, a.account_name
            ORDER BY a.account_code
        """
        return self._run_report_query("┘à┘è╪▓╪º┘å ╪º┘ä┘à╪▒╪º╪¼╪╣╪⌐", query, params, filters)

    def get_income_statement(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                COALESCE(a.account_name, '-') AS account_name,
                SUM(CASE WHEN a.nature IN ('╪»╪º╪ª┘å', 'credit', 'CREDIT') THEN COALESCE(l.credit, 0) ELSE 0 END) AS income,
                SUM(CASE WHEN a.nature IN ('┘à╪»┘è┘å', 'debit', 'DEBIT') THEN COALESCE(l.debit, 0) ELSE 0 END) AS expense,
                SUM(CASE WHEN a.nature IN ('╪»╪º╪ª┘å', 'credit', 'CREDIT') THEN COALESCE(l.credit, 0) ELSE 0 END) -
                SUM(CASE WHEN a.nature IN ('┘à╪»┘è┘å', 'debit', 'DEBIT') THEN COALESCE(l.debit, 0) ELSE 0 END) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.accounts a ON a.account_code = l.account_code
            WHERE {where_sql}
            GROUP BY a.account_name
            ORDER BY account_name
        """
        return self._run_report_query("┘é╪º╪ª┘à╪⌐ ╪º┘ä╪»╪«┘ä", query, params, filters)

    def get_cash_flow(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT
                v.v_date,
                SUM(COALESCE(l.credit, 0)) AS cash_in,
                SUM(COALESCE(l.debit, 0)) AS cash_out,
                SUM(COALESCE(l.credit, 0) - COALESCE(l.debit, 0)) AS balance
            FROM finance.vouchers v
            JOIN finance.ledger l ON l.voucher_id = v.id
            WHERE {where_sql}
            GROUP BY v.v_date
            ORDER BY v.v_date
        """
        return self._run_report_query("╪º┘ä╪¬╪»┘ü┘é╪º╪¬ ╪º┘ä┘å┘é╪»┘è╪⌐", query, params, filters)

    # ---------- Advanced ----------
    def filter_by_account(self, filters):
        if not filters.get("account_code"):
            raise ValueError("┘è╪▒╪¼┘ë ╪º╪«╪¬┘è╪º╪▒ ╪¡╪│╪º╪¿ ┘ä╪¬┘å┘ü┘è╪░ ╪º┘ä╪¬┘é╪▒┘è╪▒")
        return self.get_general_ledger(filters)

    def filter_by_heir(self, filters):
        if not filters.get("vendor_id"):
            raise ValueError("┘è╪▒╪¼┘ë ╪º╪«╪¬┘è╪º╪▒ ┘ê╪º╪▒╪½ ┘ä╪¬┘å┘ü┘è╪░ ╪º┘ä╪¬┘é╪▒┘è╪▒")
        return self.get_heir_statement(filters)

    def filter_by_property(self, filters):
        if not filters.get("property_id"):
            raise ValueError("┘è╪▒╪¼┘ë ╪º╪«╪¬┘è╪º╪▒ ╪╣┘é╪º╪▒ ┘ä╪¬┘å┘ü┘è╪░ ╪º┘ä╪¬┘é╪▒┘è╪▒")
        return self.get_property_statement(filters)

    def compare_periods(self, filters):
        start_dt = datetime.strptime(filters["date_from"], "%Y-%m-%d").date()
        end_dt = datetime.strptime(filters["date_to"], "%Y-%m-%d").date()
        days = max((end_dt - start_dt).days, 0)

        prev_end = start_dt - timedelta(days=1)
        prev_start = prev_end - timedelta(days=days)

        current_filters = dict(filters)
        previous_filters = dict(filters)

        cur_clause, cur_params = self._ledger_filter_clause(current_filters)
        prev_clause, prev_params = self._ledger_filter_clause(previous_filters)

        cur_query = f"""
            SELECT
                COALESCE(SUM(COALESCE(l.debit, 0)), 0) AS debit,
                COALESCE(SUM(COALESCE(l.credit, 0)), 0) AS credit
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            WHERE {cur_clause}
        """

        # Replace first two params (date range) for previous period
        prev_params[0] = prev_start.strftime("%Y-%m-%d")
        prev_params[1] = prev_end.strftime("%Y-%m-%d")

        prev_query = f"""
            SELECT
                COALESCE(SUM(COALESCE(l.debit, 0)), 0) AS debit,
                COALESCE(SUM(COALESCE(l.credit, 0)), 0) AS credit
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            WHERE {prev_clause}
        """

        _, cur_rows = self._run_query(cur_query, cur_params)
        _, prev_rows = self._run_query(prev_query, prev_params)

        cur_debit = float(cur_rows[0][0] or 0) if cur_rows else 0.0
        cur_credit = float(cur_rows[0][1] or 0) if cur_rows else 0.0
        prev_debit = float(prev_rows[0][0] or 0) if prev_rows else 0.0
        prev_credit = float(prev_rows[0][1] or 0) if prev_rows else 0.0

        cur_net = cur_credit - cur_debit
        prev_net = prev_credit - prev_debit

        columns = ["╪º┘ä╪¿┘å╪»", "╪º┘ä┘ü╪¬╪▒╪⌐ ╪º┘ä╪¡╪º┘ä┘è╪⌐", "╪º┘ä┘ü╪¬╪▒╪⌐ ╪º┘ä╪│╪º╪¿┘é╪⌐", "╪º┘ä┘ü╪▒┘é"]
        rows = [
            ("╪Ñ╪¼┘à╪º┘ä┘è ╪º┘ä┘à╪»┘è┘å", cur_debit, prev_debit, cur_debit - prev_debit),
            ("╪Ñ╪¼┘à╪º┘ä┘è ╪º┘ä╪»╪º╪ª┘å", cur_credit, prev_credit, cur_credit - prev_credit),
            ("╪º┘ä╪╡╪º┘ü┘è", cur_net, prev_net, cur_net - prev_net),
        ]

        summary = {
            "total_debit": cur_debit,
            "total_credit": cur_credit,
            "balance": cur_net,
        }

        return {
            "title": "┘à┘é╪º╪▒┘å╪⌐ ┘ü╪¬╪▒╪¬┘è┘å",
            "columns": columns,
            "rows": rows,
            "summary": summary,
            "date_from": filters["date_from"],
            "date_to": filters["date_to"],
        }

    # =====================================================
    # External API
    # =====================================================

    def open_report(self, report_name):
        route = self.report_to_route.get(report_name)
        if not route:
            self.notebook.select(0)
            return

        tab_key, option_name = route
        tab_index = self.tab_order.index(tab_key)
        self.notebook.select(tab_index)

        state = self.tabs[tab_key]
        state["report_var"].set(option_name)
        self._generate_report(tab_key)


def open_report(master, report_name):
    """
    Factory helper used by the main app to open a report screen directly.
    """
    screen = ReportsScreen(master)
    if report_name:
        screen.open_report(report_name)
    return screen
