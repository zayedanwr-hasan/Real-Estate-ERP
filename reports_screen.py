import os
import tkinter as tk
from datetime import date, datetime, timedelta
from decimal import Decimal
from tkinter import filedialog, messagebox, ttk
from typing import Any, Dict, Sequence

import ttkbootstrap as tb

try:
    from ttkbootstrap.widgets import DateEntry
except Exception:
    DateEntry = None

from combobox_helper import bind_searchable_combobox, set_combobox_values
from db_connection import get_connection, get_db_error_message


# =========================================================
# Export helpers
# =========================================================

def export_to_excel(data, filename):
    """
    Export a report payload to Excel using pandas/openpyxl.
    Also formats headers, freezes panes, and autosizes columns.
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise RuntimeError("مكتبة pandas/openpyxl غير مثبتة. ثبّتها أولاً.") from exc

    df = pd.DataFrame(data.get("rows", []), columns=data.get("columns", []))

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        sheet_name = (data.get("title") or "Report")[:31]
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]]

    # Header style
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze first row
    ws.freeze_panes = "A2"

    # Autosize columns
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 35)

    wb.save(filename)


def export_to_pdf(data, filename, title):
    """
    Export a report payload to PDF using reportlab.
    Supports optional Arabic shaping if arabic_reshaper + bidi are installed.
    Attempts to use a Unicode font available on the system.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        raise RuntimeError("مكتبة reportlab غير مثبتة. ثبّتها أولاً.") from exc

    def shape_text(text):
        text = "" if text is None else str(text)
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(text))
        except Exception:
            return text

    def find_unicode_font():
        candidates = [
            r"C:\Windows\Fonts\arial.ttf",
            r"C:\Windows\Fonts\segoeui.ttf",
            r"C:\Windows\Fonts\tahoma.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        ]
        for path in candidates:
            if os.path.exists(path):
                return path
        return None

    font_name = "Helvetica"
    font_path = find_unicode_font()
    if font_path:
        try:
            font_name = "AppUnicodeFont"
            pdfmetrics.registerFont(TTFont(font_name, font_path))
        except Exception:
            font_name = "Helvetica"

    doc = SimpleDocTemplate(
        filename,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="ArabicTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=16,
        leading=20,
        alignment=2,
    ))
    styles.add(ParagraphStyle(
        name="ArabicNormal",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        leading=13,
        alignment=2,
    ))
    styles.add(ParagraphStyle(
        name="ArabicHeading",
        parent=styles["Heading3"],
        fontName=font_name,
        fontSize=11,
        leading=14,
        alignment=2,
    ))

    story = []
    story.append(Paragraph(shape_text(title), styles["ArabicTitle"]))
    story.append(Spacer(1, 6))

    date_from = data.get("date_from", "")
    date_to = data.get("date_to", "")
    if date_from and date_to:
        story.append(Paragraph(shape_text(f"الفترة: {date_from} إلى {date_to}"), styles["ArabicNormal"]))
        story.append(Spacer(1, 6))

    meta_lines = data.get("meta_lines") or []
    if meta_lines:
        meta_table = Table(
            [[shape_text(label), shape_text(value)] for label, value in meta_lines if value is not None],
            colWidths=[140, 580],
        )
        meta_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ecf0f1")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#2c3e50")),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d8e0")),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(meta_table)
        story.append(Spacer(1, 8))

    columns = [shape_text(c) for c in data.get("columns", [])]
    rows = data.get("rows", [])

    def fmt(v):
        if v is None:
            return ""
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, (int, float, Decimal)):
            try:
                return f"{float(v):,.2f}"
            except Exception:
                return str(v)
        return shape_text(v)

    table_data = [columns] + [[fmt(v) for v in row] for row in rows]

    if not rows:
        story.append(Paragraph(shape_text("لا توجد بيانات لعرضها."), styles["ArabicHeading"]))
        doc.build(story)
        return

    col_count = len(columns)
    base_width = 720 / max(col_count, 1)
    col_widths = [base_width for _ in range(col_count)]

    report_table = Table(table_data, repeatRows=1, colWidths=col_widths)
    report_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d8e0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7f6")]),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(report_table)

    summary = data.get("summary", {})
    if summary:
        story.append(Spacer(1, 8))
        summary_text = " | ".join([f"{shape_text(k)}: {fmt(v)}" for k, v in summary.items()])
        story.append(Paragraph(shape_text(summary_text), styles["ArabicHeading"]))

    doc.build(story)


# =========================================================
# Main Reports Screen
# =========================================================

class ReportsScreen:
    def __init__(self, master):
        self.master = master

        self.primary_color = "#2C3E50"
        self.sidebar_color = "#34495E"
        self.accent_color = "#1ABC9C"
        self.soft_bg = "#F4F7F6"
        self.card_bg = "#FFFFFF"
        self.soft_card_bg = "#F8FAFB"
        self.summary_bg = "#EAF7F4"
        self.text_color = "#1F2D3D"

        # Match typography used in sub_coding screens.
        self.base_font = ("Segoe UI", 11)
        self.bold_font = ("Segoe UI", 11, "bold")
        self.title_font = ("Segoe UI", 20, "bold")

        self._lookup_cache = {}
        self._account_statement_dialog = None
        self._account_statement_last_options = {}
        self._report_viewer_frame = None
        self._report_viewer_tree = None
        self._report_viewer_payload = None
        self._report_viewer_font_size = 11
        self._report_viewer_footer = {}
        self._report_viewer_header = {}
        self._report_viewer_signatures = {}
        self._report_viewer_totals = {}
        self._header_frame = None
        self._viewer_user_label = "المستخدم الحالي"
        self._dashboard_container = None
        self._report_viewer_toolbar = {}
        self._report_viewer_page = 1
        self._report_viewer_page_count = 1
        self._report_viewer_zoom_percent = 100
        self._report_viewer_base_widths = {}
        self._report_viewer_paper = None

        self.report_to_route = {
            "كشف حساب وارث": ("heirs", "كشف حساب وارث"),
            "ملخص وارث": ("heirs", "ملخص وارث"),
            "أرصدة الورثة": ("heirs", "أرصدة الورثة"),
            "أرصدة الورثة او المجموعة": ("heirs", "أرصدة الورثة"),
            "توزيع الأرباح": ("heirs", "توزيع الأرباح"),
            "أرصدة الموردين/عقار": ("heirs", "أرصدة الورثة"),
            "أرصدة العملاء": ("heirs", "ملخص وارث"),
            "كشف حساب تفصيلي": ("dialog", "كشف حساب تفصيلي"),

            "كشف حساب عقار": ("properties", "كشف حساب عقار"),
            "حالة العقارات (الأراضي)": ("properties", "كشف حساب عقار"),
            "تقرير الإيرادات": ("properties", "تقرير الإيرادات"),
            "تقرير المصروفات": ("properties", "تقرير المصروفات"),
            "إيرادات ومصروفات عقار": ("properties", "صافي الربح"),
            "صافي الربح": ("properties", "صافي الربح"),

            "سندات الصرف": ("vouchers", "سندات الصرف"),
            "كشف حركة صندوق": ("vouchers", "تقرير يومي"),
            "سندات القبض": ("vouchers", "سندات القبض"),
            "تقرير يومي": ("vouchers", "تقرير يومي"),

            "دفتر الأستاذ": ("financial", "دفتر الأستاذ"),
            "دفتر اليومية العامة": ("financial", "دفتر الأستاذ"),
            "ميزان المراجعة": ("financial", "ميزان المراجعة"),
            "قائمة الدخل": ("financial", "قائمة الدخل"),
            "التدفقات النقدية": ("financial", "التدفقات النقدية"),
            "ملخص حركة الصناديق": ("financial", "التدفقات النقدية"),
            "طباعة دليل الحسابات": ("financial", "طباعة دليل الحسابات"),
        }

        self.tab_order = ["heirs", "properties", "vouchers", "financial", "advanced"]
        self.tabs = {}

        # Dynamic preview layouts keyed by logical report name.
        self.preview_layouts = {
            "account_statement": {
                "columns": ["التاريخ", "رقم السند", "البيان", "مدين", "دائن", "الرصيد"],
                "sources": ["v_date", "voucher_id", "description", "debit", "credit", "running_balance"],
                "show_totals": True,
            },
            "heirs_balances": {
                "columns": ["اسم الوارث", "الحساب", "نسبة الملكية", "المستحق", "المستلم", "المتبقي"],
                "sources": ["vendor_name", "account_name", "ownership_ratio", "debit", "credit", "balance"],
                "show_totals": True,
            },
            "property_status": {
                "columns": ["اسم الأرض", "الموقع", "المساحة", "القيمة", "الحالة (مباعة/متاحة)"],
                "sources": ["property_name", "location", "area", "balance", "status"],
                "show_totals": False,
            },
            "general_journal": {
                "columns": ["التاريخ", "رقم القيد", "الحساب", "البيان", "مدين", "دائن"],
                "sources": ["v_date", "voucher_id", "account_name", "description", "debit", "credit"],
                "show_totals": True,
            },
            "default": {
                "columns": ["التاريخ", "البيان", "مدين", "دائن", "الرصيد"],
                "sources": ["v_date", "description", "debit", "credit", "balance"],
                "show_totals": True,
            },
        }

        self._setup_styles()
        self._build_root()
        self._build_header()
        self._build_notebook()
        self._load_filter_sources()

    # =====================================================
    # UI
    # =====================================================

    def _setup_styles(self):
        style = tb.Style()
        try:
            style.theme_use(style.theme_use())
        except Exception:
            pass

        try:
            style.configure("App.Reports.TNotebook", background=self.soft_bg, borderwidth=0)
            style.configure(
                "App.Reports.TNotebook.Tab",
                font=self.bold_font,
                padding=(14, 8),
                background=self.sidebar_color,
                foreground="#ECF0F1",
            )
            style.map(
                "App.Reports.TNotebook.Tab",
                background=[("selected", self.accent_color), ("active", self.accent_color)],
                foreground=[("selected", "white"), ("active", "white")],
            )

            style.configure(
                "App.Reports.Treeview",
                rowheight=30,
                font=self.base_font,
                background="#FFFFFF",
                fieldbackground="#FFFFFF",
                bordercolor="#AEB6BF",
                borderwidth=1,
                relief="solid",
            )
            style.configure(
                "App.Reports.Treeview.Heading",
                font=self.bold_font,
                background=self.primary_color,
                foreground="#FFFFFF",
                bordercolor="#566573",
                borderwidth=1,
                relief="solid",
            )
            style.configure(
                "Viewer.Treeview",
                rowheight=30,
                font=("Segoe UI", self._report_viewer_font_size),
                background="#FFFFFF",
                fieldbackground="#FFFFFF",
                bordercolor="#000000",
                borderwidth=1,
                relief="solid",
            )
            style.configure(
                "Viewer.Treeview.Heading",
                font=("Segoe UI", 11, "bold"),
                background="#F2F2F2",
                foreground="#000000",
                bordercolor="#000000",
                borderwidth=1,
                relief="solid",
            )
        except Exception:
            pass

    def _build_root(self):
        self.frame = tk.Frame(self.master, bg=self.soft_bg)
        self.frame.pack(fill="both", expand=True)

        self.main_card = tk.Frame(
            self.frame,
            bg=self.card_bg,
            highlightthickness=1,
            highlightbackground="#D6DEE5",
        )
        self.main_card.pack(fill="both", expand=True, padx=14, pady=14)

    def _build_header(self):
        header = tk.Frame(self.main_card, bg=self.primary_color, height=58)
        self._header_frame = header
        header.pack(fill="x")
        header.pack_propagate(False)

        title = tk.Label(
            header,
            text="شاشة التقارير",
            bg=self.primary_color,
            fg="white",
            font=self.title_font,
        )
        title.pack(side="right", padx=20, pady=10)

        subtitle = tk.Label(
            header,
            text="Reports Dashboard",
            bg=self.primary_color,
            fg="#DDE6EE",
            font=self.base_font,
        )
        subtitle.pack(side="right", padx=10, pady=14)

    def _build_notebook(self):
        container = tk.Frame(self.main_card, bg=self.soft_bg)
        self._dashboard_container = container
        container.pack(fill="both", expand=True, padx=12, pady=12)

        self.notebook = ttk.Notebook(container, style="App.Reports.TNotebook")
        self.notebook.pack(fill="both", expand=True)

        self._create_report_tab(
            "heirs",
            "تقارير الورثة",
            {
                "كشف حساب وارث": self.get_heir_statement,
                "ملخص وارث": self.get_heir_summary,
                "أرصدة الورثة": self.get_heir_balances,
                "توزيع الأرباح": self.get_profit_distribution,
                "كشف حساب تفصيلي": self.get_account_statement,
                "أرصدة الموردين/عقار": self.get_heir_balances,
                "أرصدة العملاء": self.get_heir_summary,
            },
        )

        self._create_report_tab(
            "properties",
            "تقارير العقارات",
            {
                "كشف حساب عقار": self.get_property_statement,
                "حالة العقارات (الأراضي)": self.get_property_statement,
                "تقرير الإيرادات": self.get_property_income,
                "تقرير المصروفات": self.get_property_expense,
                "إيرادات ومصروفات عقار": self.get_property_profit,
                "صافي الربح": self.get_property_profit,
            },
        )

        self._create_report_tab(
            "vouchers",
            "تقارير السندات",
            {
                "سندات الصرف": self.get_payment_vouchers,
                "كشف حركة صندوق": self.get_daily_vouchers,
                "سندات القبض": self.get_receipt_vouchers,
                "تقرير يومي": self.get_daily_vouchers,
            },
        )

        self._create_report_tab(
            "financial",
            "التقارير المالية",
            {
                "دفتر الأستاذ": self.get_general_ledger,
                "دفتر اليومية العامة": self.get_general_ledger,
                "ميزان المراجعة": self.get_trial_balance,
                "قائمة الدخل": self.get_income_statement,
                "التدفقات النقدية": self.get_cash_flow,
                "ملخص حركة الصناديق": self.get_cash_flow,
                "طباعة دليل الحسابات": self.get_chart_of_accounts,
            },
        )

        self._create_report_tab(
            "advanced",
            "تقارير متقدمة",
            {
                "حسب الحساب": self.filter_by_account,
                "حسب الوارث": self.filter_by_heir,
                "حسب العقار": self.filter_by_property,
                "مقارنة فترات": self.compare_periods,
            },
        )

    def _create_report_tab(self, tab_key, tab_title, reports_map):
        tab_frame = tk.Frame(self.notebook, bg=self.soft_bg)
        self.notebook.add(tab_frame, text=tab_title)

        state = {
            "reports": reports_map,
            "report_var": tk.StringVar(value=next(iter(reports_map.keys()))),
            "date_from_var": tk.StringVar(value=self._default_date_from()),
            "date_to_var": tk.StringVar(value=date.today().strftime("%Y-%m-%d")),
            "heir_var": tk.StringVar(),
            "property_var": tk.StringVar(),
            "account_var": tk.StringVar(),
            "last_data": None,
            "tree": None,
            "summary_labels": {},
            "active_report_name": "",
            # Optional in current grid-first UX; kept for backward-compatible data hydration.
            "heir_combo": None,
            "property_combo": None,
            "account_combo": None,
        }

        table_card = self._card(tab_frame)
        table_card.pack(fill="both", expand=True, padx=10, pady=(10, 8))
        state["tree"] = self._build_results_tree(table_card)

        summary_card = self._summary_card(tab_frame)
        summary_card.pack(fill="x", padx=10, pady=(0, 10))
        state["summary_labels"] = self._build_summary_panel(summary_card)

        self.tabs[tab_key] = state

    def _card(self, parent):
        return tk.Frame(
            parent,
            bg=self.soft_card_bg,
            highlightthickness=1,
            highlightbackground="#D8E1E8",
        )

    def _summary_card(self, parent):
        return tk.Frame(
            parent,
            bg=self.summary_bg,
            highlightthickness=1,
            highlightbackground="#B8DED6",
        )

    def _field_label(self, parent, text):
        return tk.Label(
            parent,
            text=text,
            bg=self.soft_card_bg,
            fg=self.primary_color,
            font=("Segoe UI", 12, "bold"),
        )

    def _build_filter_widgets(self, parent, state):
        for c in range(5):
            parent.grid_columnconfigure(c, weight=1)

        # Row 0
        self._field_label(parent, "من تاريخ").grid(row=0, column=0, padx=6, pady=(10, 4), sticky="e")
        self._field_label(parent, "إلى تاريخ").grid(row=0, column=1, padx=6, pady=(10, 4), sticky="e")
        self._field_label(parent, "الوارث").grid(row=0, column=2, padx=6, pady=(10, 4), sticky="e")
        self._field_label(parent, "العقار").grid(row=0, column=3, padx=6, pady=(10, 4), sticky="e")
        self._field_label(parent, "الحساب").grid(row=0, column=4, padx=6, pady=(10, 4), sticky="e")

        # Row 1
        date_from_entry = ttk.Entry(parent, textvariable=state["date_from_var"], justify="center", font=self.base_font)
        date_to_entry = ttk.Entry(parent, textvariable=state["date_to_var"], justify="center", font=self.base_font)
        heir_combo = ttk.Combobox(parent, textvariable=state["heir_var"], justify="right", font=self.base_font)
        property_combo = ttk.Combobox(parent, textvariable=state["property_var"], justify="right", font=self.base_font)
        account_combo = ttk.Combobox(parent, textvariable=state["account_var"], justify="right", font=self.base_font)

        date_from_entry.grid(row=1, column=0, padx=6, pady=(0, 8), sticky="ew")
        date_to_entry.grid(row=1, column=1, padx=6, pady=(0, 8), sticky="ew")
        heir_combo.grid(row=1, column=2, padx=6, pady=(0, 8), sticky="ew")
        property_combo.grid(row=1, column=3, padx=6, pady=(0, 8), sticky="ew")
        account_combo.grid(row=1, column=4, padx=6, pady=(0, 8), sticky="ew")

        bind_searchable_combobox(heir_combo)
        bind_searchable_combobox(property_combo)
        bind_searchable_combobox(account_combo)

        state["heir_combo"] = heir_combo
        state["property_combo"] = property_combo
        state["account_combo"] = account_combo

        # Row 2
        self._field_label(parent, "نوع التقرير").grid(row=2, column=0, padx=6, pady=(6, 4), sticky="e")
        report_combo = ttk.Combobox(
            parent,
            textvariable=state["report_var"],
            state="readonly",
            justify="right",
            font=self.bold_font,
        )
        set_combobox_values(report_combo, list(state["reports"].keys()))
        report_combo.grid(row=3, column=0, columnspan=3, padx=6, pady=(0, 8), sticky="ew")

        state["report_combo"] = report_combo

    def _build_action_buttons(self, parent, tab_key):
        action_frame = tk.Frame(parent, bg=self.soft_card_bg)
        action_frame.grid(row=3, column=3, columnspan=2, padx=6, pady=(0, 8), sticky="w")

        tb.Button(
            action_frame,
            text="عرض التقرير",
            bootstyle="primary",
            command=lambda: self._generate_report(tab_key),
        ).pack(side="left", padx=4)

        tb.Button(
            action_frame,
            text="تصدير PDF",
            bootstyle="warning",
            command=lambda: self._on_export_pdf(tab_key),
        ).pack(side="left", padx=4)

        tb.Button(
            action_frame,
            text="تصدير Excel",
            bootstyle="success",
            command=lambda: self._on_export_excel(tab_key),
        ).pack(side="left", padx=4)

    def _build_results_tree(self, parent):
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        tree = ttk.Treeview(parent, show="headings", style="App.Reports.Treeview")
        y_scroll = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)

        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        tree.tag_configure("even", background="#F8FAFB")
        tree.tag_configure("odd", background="#FFFFFF")

        return tree

    def _build_summary_panel(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        row = tk.Frame(parent, bg=self.summary_bg)
        row.grid(row=0, column=0, sticky="ew", padx=6, pady=6)

        items = {}
        for key, title in (("total_debit", "إجمالي المدين"), ("total_credit", "إجمالي الدائن"), ("balance", "الرصيد الختامي")):
            frame = tk.Frame(
                row,
                bg=self.summary_bg,
                highlightthickness=1,
                highlightbackground="#B8C4CE",
                padx=8,
                pady=6,
            )
            tk.Label(frame, text=title, bg=self.summary_bg, fg=self.primary_color, font=("Segoe UI", 10, "bold"), anchor="e").pack(anchor="e")
            value_label = tk.Label(frame, text="0.00", bg=self.summary_bg, fg="#0E6251", font=("Segoe UI", 12, "bold"), anchor="e")
            value_label.pack(anchor="e")
            items[key] = {"frame": frame, "value": value_label}

        return {"row": row, "items": items}

    def _export_report(self, tab_key, extension, filetypes, exporter, success_message):
        state = self.tabs[tab_key]
        data = state.get("last_data")
        if not data or not data.get("rows"):
            messagebox.showwarning("تنبيه", "اعرض التقرير أولاً قبل التصدير")
            return

        file_name = filedialog.asksaveasfilename(
            defaultextension=extension,
            filetypes=filetypes,
        )
        if not file_name:
            return

        try:
            exporter(data, file_name)
            messagebox.showinfo("نجاح", success_message)
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def _on_export_excel(self, tab_key):
        self._export_report(
            tab_key,
            ".xlsx",
            [("Excel Files", "*.xlsx")],
            lambda data, file_name: export_to_excel(data, file_name),
            "تم تصدير التقرير إلى Excel بنجاح",
        )

    def _on_export_pdf(self, tab_key):
        self._export_report(
            tab_key,
            ".pdf",
            [("PDF Files", "*.pdf")],
            lambda data, file_name: export_to_pdf(data, file_name, data.get("title", "تقرير")),
            "تم تصدير التقرير إلى PDF بنجاح",
        )

    # =====================================================
    # Data sources
    # =====================================================

    def _default_date_from(self):
        today = date.today()
        return today.replace(day=1).strftime("%Y-%m-%d")

    def _load_filter_sources(self):
        vendors = self._fetch_lookup(
            "SELECT id, vendor_name FROM finance.vendors ORDER BY vendor_name",
            "الموردين",
        )
        properties = self._fetch_lookup(
            "SELECT id, property_name FROM finance.properties ORDER BY property_name",
            "العقارات",
        )
        accounts = self._fetch_lookup(
            "SELECT account_code, account_name FROM finance.accounts ORDER BY account_code",
            "الحسابات",
        )

        vendor_values = [f"{row[0]} - {row[1]}" for row in vendors]
        property_values = [f"{row[0]} - {row[1]}" for row in properties]
        account_values = [f"{row[0]} - {row[1]}" for row in accounts]

        self.vendor_lookup_values = vendor_values
        self.property_lookup_values = property_values
        self.account_lookup_values = account_values
        self.account_lookup_by_code = {str(row[0]): row[1] for row in accounts}
        self.property_lookup_by_id = {str(row[0]): row[1] for row in properties}

        for state in self.tabs.values():
            # Some tabs/screens currently run without filter widgets.
            heir_combo = state.get("heir_combo")
            property_combo = state.get("property_combo")
            account_combo = state.get("account_combo")
            if heir_combo is not None:
                set_combobox_values(heir_combo, vendor_values)
            if property_combo is not None:
                set_combobox_values(property_combo, property_values)
            if account_combo is not None:
                set_combobox_values(account_combo, account_values)

    def _fetch_lookup(self, query, source_name) -> Any:
        conn = get_connection()
        if not conn:
            return []
        try:
            with conn:
                with conn.cursor() as cur:
                    cur.execute(query)
                    fetched_rows = cur.fetchall()
                    if fetched_rows is None:
                        return []
                    return list(fetched_rows)
        except Exception as exc:
            messagebox.showerror("خطأ", get_db_error_message(exc, f"تعذر تحميل بيانات {source_name}"))
            return []
        finally:
            conn.close()

    def _run_query(self, query: str, params: Sequence[Any]) -> Any:
        conn = get_connection()
        if not conn:
            raise RuntimeError("تعذر الاتصال بقاعدة البيانات")

        try:
            with conn:
                with conn.cursor() as cur:
                    cur.execute(query, params)
                    fetched_rows = cur.fetchall()
                    if fetched_rows is None:
                        rows = []
                    else:
                        rows = list(fetched_rows)
                    columns = [desc[0] for desc in cur.description] if cur.description else []
            return columns, rows
        except Exception as exc:
            raise RuntimeError(get_db_error_message(exc, "خطأ في قاعدة البيانات")) from exc
        finally:
            conn.close()

    # =====================================================
    # Helpers
    # =====================================================

    def _tab_key_from_state(self, state):
        for key, value in self.tabs.items():
            if value is state:
                return key
        return "heirs"

    def _extract_prefix(self, value):
        text = (value or "").strip()
        if not text:
            return ""
        return text.split(" - ")[0].strip() if " - " in text else text

    def _is_int_like(self, value):
        return isinstance(value, int) and not isinstance(value, bool)

    def _format_cell(self, column_name, value):
        if value is None:
            return ""

        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")

        col = (column_name or "").lower()

        if self._is_int_like(value):
            if col in {"id", "voucher_id", "vouchers_count"} or col.endswith("_id") or col.endswith("_count"):
                return f"{int(value):,}"
            return f"{int(value):,}"

        if isinstance(value, (float, Decimal)):
            if col in {
                "debit", "credit", "balance", "income", "expense",
                "cash_in", "cash_out", "total_profit", "distributed_share",
                "previous_period", "current_period", "difference",
                "net_profit", "profit"
            }:
                return f"{float(value):,.2f}"
            return f"{float(value):,.2f}"

        return str(value)

    def _numeric_value(self, value):
        if value is None:
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    def _compute_summary(self, columns, rows):
        if not rows or not columns:
            return {"total_debit": 0.0, "total_credit": 0.0, "balance": 0.0}

        idx = {name.lower(): pos for pos, name in enumerate(columns)}
        total_debit = 0.0
        total_credit = 0.0

        for row in rows:
            for col_name, pos in idx.items():
                val = self._numeric_value(row[pos])
                if col_name in {"debit", "expense", "cash_out"}:
                    total_debit += val
                elif col_name in {"credit", "income", "cash_in"}:
                    total_credit += val
                elif col_name == "profit":
                    total_credit += val
                elif col_name == "distributed_share":
                    total_credit += val

        return {
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balance": total_debit - total_credit,
        }

    def _update_tree(self, state, columns, rows):
        tree = state["tree"]
        tree.delete(*tree.get_children())
        tree["columns"] = columns

        for col in columns:
            tree.heading(col, text=col, anchor="e")
            tree.column(col, anchor="e", width=155, minwidth=120, stretch=True)

        for idx, row in enumerate(rows):
            formatted = [self._format_cell(columns[i], v) for i, v in enumerate(row)]
            tag = "even" if idx % 2 == 0 else "odd"
            tree.insert("", "end", values=formatted, tags=(tag,))

    def _update_summary(self, state, summary, columns=None, show_totals=True):
        panel = state["summary_labels"]
        row = panel["row"]
        items = panel["items"]

        for child in row.winfo_children():
            child.grid_forget()

        if not show_totals:
            row.grid_remove()
            return

        row.grid()
        columns = columns or []
        count = max(len(columns), 1)
        for i in range(count):
            row.grid_columnconfigure(i, weight=1)

        debit_col = self._summary_target_column(columns, ("مدين", "debit"))
        credit_col = self._summary_target_column(columns, ("دائن", "credit"))
        balance_col = self._summary_target_column(columns, ("رصيد", "balance"))

        if debit_col is None:
            debit_col = max(count - 3, 0)
        if credit_col is None:
            credit_col = max(count - 2, 0)
        if balance_col is None:
            balance_col = max(count - 1, 0)

        items["total_debit"]["value"].configure(text=f"{summary.get('total_debit', 0.0):,.2f}")
        items["total_credit"]["value"].configure(text=f"{summary.get('total_credit', 0.0):,.2f}")
        items["balance"]["value"].configure(text=f"{summary.get('balance', 0.0):,.2f}")

        items["total_debit"]["frame"].grid(row=0, column=debit_col, padx=2, sticky="ew")
        items["total_credit"]["frame"].grid(row=0, column=credit_col, padx=2, sticky="ew")
        items["balance"]["frame"].grid(row=0, column=balance_col, padx=2, sticky="ew")

    def _resolve_layout_key(self, report_name):
        name = (report_name or "").strip()
        if name in {"كشف حساب تفصيلي", "كشف حساب"}:
            return "account_statement"
        if name in {"أرصدة الموردين/عقار", "أرصدة الورثة", "أرصدة العملاء"}:
            return "heirs_balances"
        if name in {"حالة العقارات (الأراضي)", "حالة العقارات"}:
            return "property_status"
        if name in {"دفتر اليومية العامة", "دفتر اليومية", "دفتر الأستاذ"}:
            return "general_journal"
        return "default"

    def _prepare_preview_payload(self, report_name, payload):
        layout_key = self._resolve_layout_key(report_name)
        layout: Dict[str, Any] = self.preview_layouts.get(layout_key, self.preview_layouts["default"])
        src_cols = [str(c) for c in (payload.get("columns") or [])]
        src_idx = {c.lower(): i for i, c in enumerate(src_cols)}

        def row_value(row, key):
            if key == "ownership_ratio":
                return "-"
            if key == "location":
                return "-"
            if key == "area":
                return "-"
            if key == "status":
                return "متاحة"
            if key == "account_name":
                pos = src_idx.get("account_name")
                if pos is not None and pos < len(row):
                    return row[pos]
                pos = src_idx.get("account_code")
                return row[pos] if pos is not None and pos < len(row) else "-"
            pos = src_idx.get(key.lower())
            return row[pos] if pos is not None and pos < len(row) else ""

        new_rows = []
        for row in payload.get("rows") or []:
            new_rows.append(tuple(row_value(row, key) for key in layout["sources"]))

        summary = payload.get("summary") or {"total_debit": 0.0, "total_credit": 0.0, "balance": 0.0}
        if layout_key == "account_statement":
            debit_idx = next((i for i, c in enumerate(layout["columns"]) if "مدين" in c), None)
            credit_idx = next((i for i, c in enumerate(layout["columns"]) if "دائن" in c), None)
            balance_idx = next((i for i, c in enumerate(layout["columns"]) if "رصيد" in c), None)
            details_idx = next((i for i, c in enumerate(layout["columns"]) if "البيان" in c), None)

            opening_balance = 0.0
            if new_rows and balance_idx is not None:
                first_balance = self._numeric_value(new_rows[0][balance_idx])
                first_debit = self._numeric_value(new_rows[0][debit_idx]) if debit_idx is not None else 0.0
                first_credit = self._numeric_value(new_rows[0][credit_idx]) if credit_idx is not None else 0.0
                opening_balance = first_balance - (first_debit - first_credit)

            opening_row: list[Any] = ["" for _ in layout["columns"]]
            if details_idx is not None:
                opening_row[details_idx] = "الرصيد الافتتاحي"
            if balance_idx is not None:
                opening_row[balance_idx] = opening_balance
            new_rows.insert(0, tuple(opening_row))

            final_balance = opening_balance + self._numeric_value(summary.get("balance", 0.0))
            totals_row: list[Any] = ["" for _ in layout["columns"]]
            if details_idx is not None:
                totals_row[details_idx] = "الإجمالي"
            if debit_idx is not None:
                totals_row[debit_idx] = self._numeric_value(summary.get("total_debit", 0.0))
            if credit_idx is not None:
                totals_row[credit_idx] = self._numeric_value(summary.get("total_credit", 0.0))
            if balance_idx is not None:
                totals_row[balance_idx] = final_balance
            new_rows.append(tuple(totals_row))

            summary = dict(summary)
            summary["opening_balance"] = opening_balance
            summary["final_balance"] = final_balance

        result: Dict[str, Any] = dict(payload)
        result["columns"] = list(layout["columns"])
        result["rows"] = new_rows
        result["show_totals"] = bool(layout.get("show_totals", True))
        result["summary"] = summary
        result["prepared_layout"] = layout_key
        return result

    def _ensure_report_viewer(self):
        if self._report_viewer_frame is not None and self._report_viewer_frame.winfo_exists():
            return

        frame = tk.Frame(self.main_card, bg="#444444")
        self._report_viewer_frame = frame

        toolbar = tk.Frame(frame, bg="#E5E5E5", highlightthickness=1, highlightbackground="#B8B8B8")
        toolbar.pack(side="top", fill="x")

        nav_frame = tk.Frame(toolbar, bg="#E5E5E5")
        nav_frame.pack(side="left", padx=10, pady=6)
        btn_first = tk.Button(nav_frame, text="|<", command=self._viewer_go_first, bg="#F8F8F8", fg="#222222", relief="groove", width=3)
        btn_prev = tk.Button(nav_frame, text="<", command=self._viewer_go_prev, bg="#F8F8F8", fg="#222222", relief="groove", width=3)
        page_label = tk.Label(nav_frame, text="Page 1 of 1", bg="#E5E5E5", fg="#222222", font=("Segoe UI", 10, "bold"))
        btn_next = tk.Button(nav_frame, text=">", command=self._viewer_go_next, bg="#F8F8F8", fg="#222222", relief="groove", width=3)
        btn_last = tk.Button(nav_frame, text=">|", command=self._viewer_go_last, bg="#F8F8F8", fg="#222222", relief="groove", width=3)

        btn_first.pack(side="left", padx=2)
        btn_prev.pack(side="left", padx=2)
        page_label.pack(side="left", padx=6)
        btn_next.pack(side="left", padx=2)
        btn_last.pack(side="left", padx=2)

        action_frame = tk.Frame(toolbar, bg="#E5E5E5")
        action_frame.pack(side="right", padx=10, pady=6)

        tk.Button(action_frame, text="رجوع", command=self._back_to_reports_dashboard, bg="#F8F8F8", fg="#1F2D3D", relief="groove").pack(side="right", padx=3)
        tk.Button(action_frame, text="Export Excel", command=self._viewer_export_excel, bg="#F8F8F8", fg="#1F2D3D", relief="groove").pack(side="right", padx=3)
        tk.Button(action_frame, text="Export PDF", command=self._viewer_export_pdf, bg="#F8F8F8", fg="#1F2D3D", relief="groove").pack(side="right", padx=3)
        tk.Button(action_frame, text="طباعة", command=self._viewer_direct_print, bg="#F8F8F8", fg="#1F2D3D", relief="groove").pack(side="right", padx=3)

        zoom_combo = ttk.Combobox(action_frame, values=["75%", "100%", "125%", "150%"], state="readonly", width=8, justify="center")
        zoom_combo.set("100%")
        zoom_combo.bind("<<ComboboxSelected>>", lambda e: self._set_viewer_zoom(zoom_combo.get()))
        zoom_combo.pack(side="right", padx=3)
        tk.Label(action_frame, text="Zoom", bg="#E5E5E5", fg="#1F2D3D").pack(side="right", padx=(6, 0))

        stage = tk.Frame(frame, bg="#444444")
        stage.pack(side="top", fill="both", expand=True)
        stage.grid_columnconfigure(0, weight=1)
        stage.grid_rowconfigure(0, weight=1)

        paper = tk.Frame(stage, bg="white", width=960, highlightthickness=1, highlightbackground="#9A9A9A")
        paper.grid(row=0, column=0, sticky="ns", padx=20, pady=14)
        self._report_viewer_paper = paper

        report_header = tk.Frame(paper, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
        report_header.pack(side="top", fill="x", padx=12, pady=(12, 8))
        report_header.grid_columnconfigure(0, weight=1)
        report_header.grid_columnconfigure(1, weight=2)
        report_header.grid_columnconfigure(2, weight=1)

        company_lbl = tk.Label(
            report_header,
            text="اسم الشركة\nرقم الهاتف / البريد الإلكتروني",
            bg="#FFFFFF",
            fg="#1F2D3D",
            font=("Segoe UI", 10, "bold"),
            anchor="e",
            justify="right",
        )
        company_lbl.grid(row=0, column=2, sticky="ew", padx=8, pady=(8, 4))

        title_lbl = tk.Label(
            report_header,
            text="تقرير",
            bg="#FFFFFF",
            fg="#8B0000",
            font=("Segoe UI", 14, "bold"),
            anchor="center",
            justify="center",
            pady=6,
        )
        title_lbl.grid(row=0, column=1, sticky="ew", padx=8, pady=(8, 4))

        left_info_lbl = tk.Label(
            report_header,
            text="",
            bg="#FFFFFF",
            fg="#1F2D3D",
            font=("Segoe UI", 10, "bold"),
            anchor="e",
            justify="right",
        )
        left_info_lbl.grid(row=0, column=0, sticky="e", padx=8, pady=(8, 4))

        meta_frame = tk.Frame(report_header, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
        meta_frame.grid(row=1, column=0, columnspan=3, sticky="ew", padx=8, pady=(0, 8))
        meta_frame.grid_columnconfigure(0, weight=1)

        meta_lbl = tk.Label(meta_frame, text="-", bg="#FFFFFF", fg="#1F2D3D", font=("Segoe UI", 10), anchor="e", justify="right")
        meta_lbl.grid(row=0, column=0, sticky="ew", padx=6, pady=6)

        grid_frame = tk.Frame(paper, bg="white")
        grid_frame.pack(side="top", fill="both", expand=True, padx=12, pady=(0, 8))
        grid_frame.grid_rowconfigure(0, weight=1)
        grid_frame.grid_columnconfigure(0, weight=1)

        tree = ttk.Treeview(grid_frame, show="headings", style="Viewer.Treeview")
        tree.tag_configure("opening", background="#FFFFFF", foreground="#000000")
        tree.tag_configure("totals", background="#F7F7F7", foreground="#000000")
        tree.tag_configure("even", background="#FFFFFF", foreground="#000000")
        tree.tag_configure("odd", background="#FCFCFC", foreground="#000000")

        y_scroll = ttk.Scrollbar(grid_frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(grid_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        totals_frame = tk.Frame(paper, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
        totals_frame.pack(side="top", fill="x", padx=12, pady=(0, 8))
        total_debit_lbl = tk.Label(totals_frame, text="إجمالي المدين: 0.00", bg="#FFFFFF", fg="#000000", font=("Segoe UI", 10, "bold"), anchor="e")
        total_credit_lbl = tk.Label(totals_frame, text="إجمالي الدائن: 0.00", bg="#FFFFFF", fg="#000000", font=("Segoe UI", 10, "bold"), anchor="e")
        balance_lbl = tk.Label(totals_frame, text="الرصيد الختامي: 0.00", bg="#FFFFFF", fg="#000000", font=("Segoe UI", 10, "bold"), anchor="e")

        footer = tk.Frame(paper, bg="#FFFFFF", highlightthickness=1, highlightbackground="#000000")
        footer.pack(side="bottom", fill="x", padx=12, pady=(0, 12))

        signatures_row = tk.Frame(footer, bg="#FFFFFF")
        signatures_row.pack(fill="x", padx=8, pady=(16, 4))
        for i, title in enumerate(("أمين الصندوق", "المحاسب", "المدير")):
            signatures_row.grid_columnconfigure(i, weight=1)
            tk.Label(
                signatures_row,
                text=title,
                bg="#FFFFFF",
                fg="#1F2D3D",
                font=("Segoe UI", 10, "bold"),
                highlightthickness=1,
                highlightbackground="#000000",
                pady=8,
                anchor="center",
            ).grid(row=0, column=i, padx=4, sticky="ew")

        legal_lbl = tk.Label(
            footer,
            text="باستثناء الخطأ والسهو يعتبر هذا الكشف صحيحاً ما لم يرد اعتراض خطي.",
            bg="#FFFFFF",
            fg="#1F2D3D",
            font=("Segoe UI", 9),
            anchor="center",
            justify="center",
        )
        legal_lbl.pack(fill="x", padx=8, pady=(2, 8))

        self._report_viewer_tree = tree
        self._report_viewer_toolbar = {
            "page": page_label,
            "first": btn_first,
            "prev": btn_prev,
            "next": btn_next,
            "last": btn_last,
            "zoom": zoom_combo,
        }
        self._report_viewer_footer = {"frame": footer}
        self._report_viewer_signatures = {"legal": legal_lbl}
        self._report_viewer_header = {
            "title": title_lbl,
            "company": company_lbl,
            "left_info": left_info_lbl,
            "meta": meta_lbl,
        }
        self._report_viewer_totals = {
            "frame": totals_frame,
            "debit": total_debit_lbl,
            "credit": total_credit_lbl,
            "balance": balance_lbl,
        }

    def _show_statement_viewer_shell(self):
        return

    def _show_report_viewer(self):
        self._ensure_report_viewer()

        if self._account_statement_dialog is not None and self._account_statement_dialog.winfo_exists():
            try:
                self._account_statement_dialog.grab_release()
            except Exception:
                pass
            self._account_statement_dialog.destroy()
            self._account_statement_dialog = None

        if self._header_frame is not None and self._header_frame.winfo_exists() and self._header_frame.winfo_ismapped():
            self._header_frame.pack_forget()

        dashboard = self._dashboard_container
        if dashboard is not None and dashboard.winfo_exists() and dashboard.winfo_ismapped():
            dashboard.pack_forget()

        frame = self._report_viewer_frame
        if frame is not None and frame.winfo_exists() and not frame.winfo_ismapped():
            frame.pack(fill="both", expand=True)

        if os.name == "nt":
            top = self.master.winfo_toplevel()
            try:
                top.state("zoomed")
            except Exception:
                pass

    def _back_to_reports_dashboard(self):
        frame = self._report_viewer_frame
        if frame is not None and frame.winfo_exists() and frame.winfo_ismapped():
            frame.pack_forget()

        if self._header_frame is not None and self._header_frame.winfo_exists() and not self._header_frame.winfo_ismapped():
            self._header_frame.pack(fill="x")

        dashboard = self._dashboard_container
        if dashboard is not None and dashboard.winfo_exists() and not dashboard.winfo_ismapped():
            dashboard.pack(fill="both", expand=True, padx=12, pady=12)

    def _render_report_viewer(self, report_name, payload):
        self._show_report_viewer()
        self._reset_report_viewer()

        viewer_payload = payload if payload.get("prepared_layout") else self._prepare_preview_payload(report_name, payload)
        self._report_viewer_payload = viewer_payload
        self._report_viewer_page = 1
        self._report_viewer_page_count = 1

        tree = self._report_viewer_tree
        if tree is None or not tree.winfo_exists():
            return

        columns = viewer_payload.get("columns", [])
        rows = viewer_payload.get("rows", [])
        tree.delete(*tree.get_children())
        tree["columns"] = columns

        self._report_viewer_base_widths = {}
        zoom_scale = self._report_viewer_zoom_percent / 100.0
        for col in columns:
            base_width = 280 if "البيان" in col else (200 if col in {"الحساب", "اسم الحساب"} else 150)
            width = int(base_width * zoom_scale)
            self._report_viewer_base_widths[col] = base_width
            tree.heading(col, text=col, anchor="e")
            tree.column(col, anchor="e", width=width, minwidth=110, stretch=True)

        for idx, row in enumerate(rows):
            formatted = [self._format_cell(columns[i], row[i]) if i < len(row) else "" for i in range(len(columns))]
            tag = "even" if idx % 2 == 0 else "odd"
            if any(str(cell) == "الرصيد الافتتاحي" for cell in formatted):
                tag = "opening"
            elif any(str(cell) == "الإجمالي" for cell in formatted):
                tag = "totals"
            tree.insert("", "end", values=formatted, tags=(tag,))

        tree.yview_moveto(0.0)
        self._set_report_viewer_header(report_name, viewer_payload)
        self._update_viewer_footer(viewer_payload.get("summary", {}), columns, viewer_payload.get("show_totals", True))
        self._update_viewer_page_indicator()

    def _update_viewer_footer(self, summary, columns, show_totals):
        totals = self._report_viewer_totals
        frame = totals.get("frame")
        if not frame:
            return

        for child in frame.winfo_children():
            child.grid_forget()

        if not show_totals:
            frame.pack_forget()
            return

        if not frame.winfo_ismapped():
            frame.pack(side="top", fill="x", padx=12, pady=(0, 8))

        columns = columns or []
        count = max(1, len(columns))
        for i in range(count):
            frame.grid_columnconfigure(i, weight=1)

        debit_col = self._summary_target_column(columns, ("مدين", "debit"))
        credit_col = self._summary_target_column(columns, ("دائن", "credit"))
        balance_col = self._summary_target_column(columns, ("رصيد", "balance"))

        if debit_col is None:
            debit_col = max(count - 3, 0)
        if credit_col is None:
            credit_col = max(count - 2, 0)
        if balance_col is None:
            balance_col = max(count - 1, 0)

        final_balance = summary.get("final_balance", summary.get("balance", 0.0))
        totals["debit"].configure(text=f"إجمالي المدين: {summary.get('total_debit', 0.0):,.2f}")
        totals["credit"].configure(text=f"إجمالي الدائن: {summary.get('total_credit', 0.0):,.2f}")
        totals["balance"].configure(text=f"الرصيد الختامي: {self._numeric_value(final_balance):,.2f}")

        totals["debit"].grid(row=0, column=debit_col, sticky="e", padx=6, pady=6)
        totals["credit"].grid(row=0, column=credit_col, sticky="e", padx=6, pady=6)
        totals["balance"].grid(row=0, column=balance_col, sticky="e", padx=6, pady=6)

    def _set_report_viewer_header(self, report_name, payload):
        if not self._report_viewer_header:
            return

        title_lbl = self._report_viewer_header.get("title")
        if title_lbl is not None:
            title_lbl.configure(text=report_name or payload.get("title") or "تقرير")

        meta_lines = payload.get("meta_lines") or []
        meta_text_parts = []
        date_from = payload.get("date_from", "-") or "-"
        date_to = payload.get("date_to", "-") or "-"
        meta_text_parts.append(f"الفترة: {date_from} -> {date_to}")
        for key, value in meta_lines:
            meta_text_parts.append(f"{key}: {value}")

        meta_lbl = self._report_viewer_header.get("meta")
        if meta_lbl is not None:
            meta_lbl.configure(text=" | ".join(meta_text_parts) if meta_text_parts else "-")

        left_info_lbl = self._report_viewer_header.get("left_info")
        if left_info_lbl is not None:
            now = datetime.now()
            left_info_lbl.configure(
                text=(
                    f"التاريخ: {now.strftime('%Y-%m-%d')}\n"
                    f"الوقت: {now.strftime('%H:%M:%S')}\n"
                    f"الصفحة: {self._report_viewer_page}/{self._report_viewer_page_count}\n"
                    f"المستخدم: {self._viewer_user_label}"
                )
            )

        legal_lbl = self._report_viewer_signatures.get("legal")
        if legal_lbl is not None:
            legal_lbl.configure(text="باستثناء الخطأ والسهو يعتبر هذا الكشف صحيحاً ما لم يرد اعتراض خطي.")

    def _summary_target_column(self, columns, candidates):
        for idx, col in enumerate(columns or []):
            low = str(col).lower()
            if any(token in low for token in candidates):
                return idx
        return None

    def _present_payload(self, tab_key, payload, filters, report_name):
        state = self.tabs[tab_key]
        payload = self._normalize_payload(payload, filters, report_name)
        payload = self._prepare_preview_payload(report_name, payload)
        state["active_report_name"] = report_name

        rows = payload.get("rows", [])
        cols = payload.get("columns", [])
        self._update_tree(state, cols, rows)
        self._update_summary(state, payload.get("summary", {}), cols, payload.get("show_totals", True))
        state["last_data"] = payload
        return payload, bool(rows)

    def _extract_optional_int(self, value, field_name):
        token = self._extract_prefix(value)
        if not token:
            return ""
        if token.isdigit():
            return int(token)
        raise ValueError(f"يرجى اختيار {field_name} من القائمة")

    def _reset_report_viewer(self):
        tree = self._report_viewer_tree
        if tree is not None and tree.winfo_exists():
            tree.delete(*tree.get_children())
            tree["columns"] = ()

        self._report_viewer_payload = None
        self._report_viewer_page = 1
        self._report_viewer_page_count = 1
        self._report_viewer_base_widths = {}

        title_lbl = self._report_viewer_header.get("title")
        if title_lbl is not None:
            title_lbl.configure(text="تقرير")

        meta_lbl = self._report_viewer_header.get("meta")
        if meta_lbl is not None:
            meta_lbl.configure(text="-")

        left_info_lbl = self._report_viewer_header.get("left_info")
        if left_info_lbl is not None:
            left_info_lbl.configure(text="")

        self._update_viewer_footer({"total_debit": 0.0, "total_credit": 0.0, "balance": 0.0}, [], True)
        self._update_viewer_page_indicator()

    def _update_viewer_page_indicator(self):
        page_lbl = self._report_viewer_toolbar.get("page")
        if page_lbl is not None:
            page_lbl.configure(text=f"Page {self._report_viewer_page} of {self._report_viewer_page_count}")

        only_page = self._report_viewer_page_count <= 1
        for key in ("first", "prev", "next", "last"):
            btn = self._report_viewer_toolbar.get(key)
            if btn is not None:
                btn.configure(state="disabled" if only_page else "normal")

        # Keep page/time/user block in header current after page changes.
        self._set_report_viewer_header(None, self._report_viewer_payload or {})

    def _viewer_go_first(self):
        self._report_viewer_page = 1
        self._update_viewer_page_indicator()

    def _viewer_go_prev(self):
        self._report_viewer_page = max(1, self._report_viewer_page - 1)
        self._update_viewer_page_indicator()

    def _viewer_go_next(self):
        self._report_viewer_page = min(self._report_viewer_page_count, self._report_viewer_page + 1)
        self._update_viewer_page_indicator()

    def _viewer_go_last(self):
        self._report_viewer_page = self._report_viewer_page_count
        self._update_viewer_page_indicator()

    def _set_viewer_zoom(self, zoom_text):
        text = str(zoom_text or "100%").strip().replace("%", "")
        try:
            percent = int(text)
        except Exception:
            percent = 100
        percent = max(75, min(150, percent))
        self._report_viewer_zoom_percent = percent

        scale = percent / 100.0
        self._report_viewer_font_size = max(9, int(round(11 * scale)))
        row_height = max(24, int(round(30 * scale)))
        tb.Style().configure("Viewer.Treeview", font=("Segoe UI", self._report_viewer_font_size), rowheight=row_height)

        tree = self._report_viewer_tree
        if tree is not None and tree.winfo_exists():
            for col in tree["columns"]:
                base = self._report_viewer_base_widths.get(col, 150)
                tree.column(col, width=int(base * scale))

        if self._report_viewer_paper is not None and self._report_viewer_paper.winfo_exists():
            paper_width = int(round(960 * scale))
            paper_width = max(900, min(1100, paper_width))
            self._report_viewer_paper.configure(width=paper_width)

    def _viewer_zoom(self, step):
        current = self._report_viewer_zoom_percent + (25 * step)
        current = max(75, min(150, current))
        value = f"{current}%"
        zoom_combo = self._report_viewer_toolbar.get("zoom")
        if zoom_combo is not None:
            zoom_combo.set(value)
        self._set_viewer_zoom(value)

    def _viewer_export_pdf(self):
        data = self._report_viewer_payload
        if not data:
            messagebox.showwarning("تنبيه", "لا توجد بيانات للتصدير")
            return
        file_name = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
        if file_name:
            export_to_pdf(data, file_name, data.get("title", "تقرير"))

    def _viewer_export_excel(self):
        data = self._report_viewer_payload
        if not data:
            messagebox.showwarning("تنبيه", "لا توجد بيانات للتصدير")
            return
        file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if file_name:
            export_to_excel(data, file_name)

    def _viewer_direct_print(self):
        data = self._report_viewer_payload
        if not data:
            messagebox.showwarning("تنبيه", "لا توجد بيانات للطباعة")
            return
        try:
            import tempfile
            temp_file = os.path.join(tempfile.gettempdir(), "landledger_report_print.pdf")
            export_to_pdf(data, temp_file, data.get("title", "تقرير"))
            if os.name == "nt":
                os.startfile(temp_file)
            else:
                messagebox.showinfo("تم", f"تم إنشاء ملف الطباعة: {temp_file}")
        except Exception as exc:
            messagebox.showerror("خطأ", f"تعذر تنفيذ الطباعة: {exc}")

    def _collect_account_statement_filters(self, state):
        date_from = self._read_date_widget(state["date_from_widget"])
        date_to = self._read_date_widget(state["date_to_widget"])

        try:
            from_dt = datetime.strptime(date_from, "%Y-%m-%d").date()
            to_dt = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            messagebox.showwarning("تنبيه", "صيغة التاريخ يجب أن تكون YYYY-MM-DD")
            return None

        if from_dt > to_dt:
            messagebox.showwarning("تنبيه", "تاريخ البداية يجب أن يكون قبل تاريخ النهاية")
            return None

        all_accounts = bool(state["all_accounts_var"].get())
        include_unposted = bool(state["include_unposted_var"].get())
        account_value = state["account_var"].get().strip()
        account_code = "" if all_accounts else self._extract_prefix(account_value)

        if not all_accounts and not account_code:
            messagebox.showwarning("تنبيه", "يرجى اختيار حساب لتنفيذ التقرير")
            return None

        account_label = "جميع الحسابات" if all_accounts else self._lookup_display_by_prefix(self.account_lookup_values, account_code)
        filters = {
            "date_from": date_from,
            "date_to": date_to,
            "account_code": account_code,
            "account_label": account_label or account_code,
            "all_accounts": all_accounts,
            "include_unposted": include_unposted,
        }
        self._account_statement_last_options = dict(filters)
        return filters

    def _open_account_statement_dialog(self):
        dialog = self._account_statement_dialog
        if dialog and dialog.winfo_exists():
            dialog.deiconify()
            dialog.lift()
            dialog.focus_force()
            return

        owner = self.master.winfo_toplevel()
        dialog = tk.Toplevel(owner)
        self._account_statement_dialog = dialog
        dialog.title("خيارات كشف الحساب")
        dialog.configure(bg=self.sidebar_color)
        dialog.resizable(False, False)
        dialog.transient(owner)
        dialog.grab_set()

        width, height = 650, 450
        sw, sh = dialog.winfo_screenwidth(), dialog.winfo_screenheight()
        dialog.geometry(f"{width}x{height}+{(sw - width) // 2}+{(sh - height) // 2}")

        def close_dialog():
            self._account_statement_dialog = None
            try:
                dialog.grab_release()
            except Exception:
                pass
            dialog.destroy()

        dialog.protocol("WM_DELETE_WINDOW", close_dialog)

        defaults = self._account_statement_last_options or {}
        date_from_value = defaults.get("date_from", self._default_date_from())
        date_to_value = defaults.get("date_to", date.today().strftime("%Y-%m-%d"))
        account_value = self._lookup_display_by_prefix(self.account_lookup_values, defaults.get("account_code", ""))

        header = tk.Frame(dialog, bg=self.primary_color, height=54)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="خيارات كشف الحساب", bg=self.primary_color, fg="white", font=("Segoe UI", 15, "bold"), anchor="e").pack(fill="both", padx=16)

        footer = tk.Frame(dialog, bg=self.sidebar_color)
        footer.pack(side="bottom", fill="x", padx=15, pady=15)

        body = tk.Frame(dialog, bg=self.sidebar_color)
        body.pack(fill="both", expand=True, padx=15, pady=(15, 8))

        content = tk.Frame(body, bg=self.sidebar_color)
        content.pack(fill="both", expand=True)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)

        state: Dict[str, Any] = {
            "date_from_widget": None,
            "date_to_widget": None,
            "account_var": tk.StringVar(value=account_value),
            "all_accounts_var": tk.BooleanVar(value=bool(defaults.get("all_accounts", False))),
            "include_unposted_var": tk.BooleanVar(value=bool(defaults.get("include_unposted", False))),
            "account_combo": None,
        }
        dialog._state = state

        def make_label(text, row, col):
            tk.Label(content, text=text, bg=self.sidebar_color, fg="white", font=("Segoe UI", 10, "bold"), anchor="e").grid(row=row, column=col, padx=15, pady=(0, 10), sticky="ew")

        make_label("من تاريخ", 0, 1)
        make_label("إلى تاريخ", 0, 0)
        make_label("الحساب", 2, 1)

        date_from_widget = self._build_date_widget(content, date_from_value)
        date_to_widget = self._build_date_widget(content, date_to_value)
        account_combo = ttk.Combobox(content, textvariable=state["account_var"], justify="right", font=self.base_font)
        set_combobox_values(account_combo, self.account_lookup_values)
        bind_searchable_combobox(account_combo)

        state["date_from_widget"] = date_from_widget
        state["date_to_widget"] = date_to_widget
        state["account_combo"] = account_combo

        date_from_widget.grid(row=1, column=1, padx=15, pady=(0, 10), sticky="ew")
        date_to_widget.grid(row=1, column=0, padx=15, pady=(0, 10), sticky="ew")
        account_combo.grid(row=3, column=0, columnspan=2, padx=15, pady=(0, 10), sticky="ew")

        checks = tk.Frame(content, bg=self.sidebar_color)
        checks.grid(row=4, column=0, columnspan=2, padx=15, pady=(6, 12), sticky="ew")
        checks.grid_columnconfigure(0, weight=1)
        checks.grid_columnconfigure(1, weight=1)

        tk.Checkbutton(checks, text="كل الحسابات", variable=state["all_accounts_var"], command=lambda: self._toggle_account_combo_state(state), bg=self.sidebar_color, fg="white", activebackground=self.sidebar_color, activeforeground="white", selectcolor=self.primary_color).grid(row=0, column=1, sticky="e")
        tk.Checkbutton(checks, text="إدراج غير المرحلة", variable=state["include_unposted_var"], bg=self.sidebar_color, fg="white", activebackground=self.sidebar_color, activeforeground="white", selectcolor=self.primary_color).grid(row=0, column=0, sticky="w")

        self._toggle_account_combo_state(state)

        tk.Button(footer, text="إلغاء", command=close_dialog, bg="#dc3545", fg="white", relief="flat", padx=20, pady=8, font=("Segoe UI", 10, "bold")).pack(side="right", padx=6)
        tk.Button(footer, text="طباعة", command=lambda: self._submit_account_statement(dialog, "print"), bg="#6f42c1", fg="white", relief="flat", padx=20, pady=8, font=("Segoe UI", 10, "bold")).pack(side="right", padx=6)
        tk.Button(footer, text="عرض التقرير", command=lambda: self._submit_account_statement(dialog, "view"), bg="#0d6efd", fg="white", relief="flat", padx=20, pady=8, font=("Segoe UI", 10, "bold")).pack(side="right", padx=6)

    def _submit_account_statement(self, dialog, action):
        state = getattr(dialog, "_state", None)
        if not state:
            return
        filters = self._collect_account_statement_filters(state)
        if not filters:
            return

        try:
            payload = self.get_account_statement(filters)
            payload = self._normalize_payload(payload, filters, "كشف حساب تفصيلي")
            payload["meta_lines"] = [
                ("الحساب", filters["account_label"]),
                ("كل الحسابات", "نعم" if filters["all_accounts"] else "لا"),
                ("إدراج غير المرحلة", "نعم" if filters["include_unposted"] else "لا"),
            ]
        except Exception as exc:
            messagebox.showerror("خطأ", f"تعذر إنشاء التقرير: {exc}")
            return

        try:
            dialog.grab_release()
        except Exception:
            pass
        dialog.destroy()
        self._account_statement_dialog = None

        if action in {"print", "preview", "view"}:
            self._render_report_viewer("كشف حساب تفصيلي", payload)
            if action == "print":
                self._viewer_direct_print()

    def get_heir_statement(self, filters):
        where_sql, params = self._ledger_filter_clause(filters, property_field="vd.property_id")
        query = f"""
            SELECT v.id AS voucher_id, v.v_date, COALESCE(vd.vendor_name, '-') AS vendor_name,
                   COALESCE(p.property_name, '-') AS property_name, COALESCE(a.account_name, '-') AS account_name,
                   COALESCE(l.debit, 0) AS debit, COALESCE(l.credit, 0) AS credit,
                   COALESCE(l.debit, 0) - COALESCE(l.credit, 0) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.vendors vd ON vd.id = l.vendor_id
            LEFT JOIN finance.properties p ON p.id = vd.property_id
            LEFT JOIN finance.accounts a ON a.account_code = l.account_code
            WHERE {where_sql}
            ORDER BY v.v_date, v.id
        """
        return self._run_report_query("كشف حساب وارث", query, params, filters)

    def get_heir_summary(self, filters):
        where_sql, params = self._ledger_filter_clause(filters, property_field="vd.property_id")
        query = f"""
            SELECT COALESCE(vd.vendor_name, 'غير محدد') AS vendor_name,
                   SUM(COALESCE(l.debit, 0)) AS debit,
                   SUM(COALESCE(l.credit, 0)) AS credit,
                   SUM(COALESCE(l.debit, 0) - COALESCE(l.credit, 0)) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.vendors vd ON vd.id = l.vendor_id
            WHERE {where_sql}
            GROUP BY vd.vendor_name
            ORDER BY vendor_name
        """
        return self._run_report_query("ملخص وارث", query, params, filters)

    def get_heir_balances(self, filters):
        return self.get_heir_summary(filters)

    def get_profit_distribution(self, filters):
        return self.get_heir_summary(filters)

    def get_property_statement(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT v.id AS voucher_id, v.v_date, COALESCE(p.property_name, '-') AS property_name,
                   COALESCE(a.account_name, '-') AS account_name,
                   COALESCE(l.debit, 0) AS debit, COALESCE(l.credit, 0) AS credit,
                   COALESCE(l.credit, 0) - COALESCE(l.debit, 0) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.properties p ON p.id = l.property_id
            LEFT JOIN finance.accounts a ON a.account_code = l.account_code
            WHERE {where_sql}
            ORDER BY v.v_date, v.id
        """
        return self._run_report_query("كشف حساب عقار", query, params, filters)

    def get_property_income(self, filters):
        return self.get_property_statement(filters)

    def get_property_expense(self, filters):
        return self.get_property_statement(filters)

    def get_property_profit(self, filters):
        return self.get_property_statement(filters)

    def get_payment_vouchers(self, filters):
        return self.get_property_statement(filters)

    def get_receipt_vouchers(self, filters):
        return self.get_property_statement(filters)

    def get_daily_vouchers(self, filters):
        return self.get_property_statement(filters)

    def get_general_ledger(self, filters):
        where_sql, params = self._ledger_filter_clause(filters)
        query = f"""
            SELECT v.v_date, v.id AS voucher_id, COALESCE(a.account_name, '-') AS account_name,
                   COALESCE(l.debit, 0) AS debit, COALESCE(l.credit, 0) AS credit,
                   COALESCE(l.debit, 0) - COALESCE(l.credit, 0) AS balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            LEFT JOIN finance.accounts a ON a.account_code = l.account_code
            WHERE {where_sql}
            ORDER BY v.v_date, v.id
        """
        return self._run_report_query("دفتر الأستاذ", query, params, filters)

    def get_trial_balance(self, filters):
        return self.get_general_ledger(filters)

    def get_income_statement(self, filters):
        return self.get_general_ledger(filters)

    def get_cash_flow(self, filters):
        return self.get_general_ledger(filters)

    def get_chart_of_accounts(self, filters):
        columns, rows = self._run_query("SELECT account_code, account_name, COALESCE(nature, '-') AS nature FROM finance.accounts ORDER BY account_code", [])
        return {
            "title": "طباعة دليل الحسابات",
            "columns": columns,
            "rows": rows,
            "summary": {"total_debit": 0.0, "total_credit": 0.0, "balance": 0.0},
            "date_from": filters.get("date_from", ""),
            "date_to": filters.get("date_to", ""),
        }

    def get_account_statement(self, filters):
        include_account = not filters.get("all_accounts")
        where_sql, params = self._ledger_filter_clause(filters, include_account=include_account)
        where_sql += self._posted_clause(filters.get("include_unposted"))
        query = f"""
            SELECT v.v_date, v.id AS voucher_id, COALESCE(v.description, '-') AS description,
                   COALESCE(l.debit, 0) AS debit, COALESCE(l.credit, 0) AS credit,
                   SUM(COALESCE(l.debit, 0) - COALESCE(l.credit, 0)) OVER (ORDER BY v.v_date, v.id, l.id) AS running_balance
            FROM finance.ledger l
            JOIN finance.vouchers v ON v.id = l.voucher_id
            WHERE {where_sql}
            ORDER BY v.v_date, v.id, l.id
        """
        payload = self._run_report_query("كشف حساب تفصيلي", query, params, filters)
        payload["meta_lines"] = [
            ("الحساب", filters.get("account_label") or ("جميع الحسابات" if filters.get("all_accounts") else filters.get("account_code", ""))),
            ("المرحلة", "تشمل غير المرحلة" if filters.get("include_unposted") else "المرحلة فقط"),
        ]
        return payload

    def filter_by_account(self, filters):
        if not filters.get("account_code"):
            raise ValueError("يرجى اختيار حساب لتنفيذ التقرير")
        return self.get_general_ledger(filters)

    def filter_by_heir(self, filters):
        if not filters.get("vendor_id"):
            raise ValueError("يرجى اختيار وارث لتنفيذ التقرير")
        return self.get_heir_statement(filters)

    def filter_by_property(self, filters):
        if not filters.get("property_id"):
            raise ValueError("يرجى اختيار عقار لتنفيذ التقرير")
        return self.get_property_statement(filters)

    def compare_periods(self, filters):
        start_dt = datetime.strptime(filters["date_from"], "%Y-%m-%d").date()
        end_dt = datetime.strptime(filters["date_to"], "%Y-%m-%d").date()
        days = max((end_dt - start_dt).days, 0)
        prev_end = start_dt - timedelta(days=1)
        prev_start = prev_end - timedelta(days=days)

        current = self.get_general_ledger(filters)
        prev_filters = dict(filters)
        prev_filters["date_from"] = prev_start.strftime("%Y-%m-%d")
        prev_filters["date_to"] = prev_end.strftime("%Y-%m-%d")
        previous = self.get_general_ledger(prev_filters)

        cur = current.get("summary", {})
        prv = previous.get("summary", {})
        rows = [
            ("إجمالي المدين", cur.get("total_debit", 0.0), prv.get("total_debit", 0.0), cur.get("total_debit", 0.0) - prv.get("total_debit", 0.0)),
            ("إجمالي الدائن", cur.get("total_credit", 0.0), prv.get("total_credit", 0.0), cur.get("total_credit", 0.0) - prv.get("total_credit", 0.0)),
            ("الصافي", cur.get("balance", 0.0), prv.get("balance", 0.0), cur.get("balance", 0.0) - prv.get("balance", 0.0)),
        ]
        return {
            "title": "مقارنة فترتين",
            "columns": ["البند", "الفترة الحالية", "الفترة السابقة", "الفرق"],
            "rows": rows,
            "summary": cur,
            "date_from": filters.get("date_from", ""),
            "date_to": filters.get("date_to", ""),
        }

    def _validate_filters(self, tab_key):
        state = self.tabs[tab_key]
        date_from = state["date_from_var"].get().strip()
        date_to = state["date_to_var"].get().strip()

        try:
            from_dt = datetime.strptime(date_from, "%Y-%m-%d").date()
            to_dt = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            messagebox.showwarning("تنبيه", "صيغة التاريخ يجب أن تكون YYYY-MM-DD")
            return None

        if from_dt > to_dt:
            messagebox.showwarning("تنبيه", "تاريخ البداية يجب أن يكون قبل تاريخ النهاية")
            return None

        try:
            vendor_id = self._extract_optional_int(state["heir_var"].get(), "الوارث")
            property_id = self._extract_optional_int(state["property_var"].get(), "العقار")
        except ValueError as exc:
            messagebox.showwarning("تنبيه", str(exc))
            return None

        return {
            "date_from": date_from,
            "date_to": date_to,
            "vendor_id": vendor_id,
            "property_id": property_id,
            "account_code": self._extract_prefix(state["account_var"].get()),
        }

    def _ledger_filter_clause(self, filters, date_field="v.v_date", property_field="l.property_id", account_field="l.account_code", include_account=True):
        conditions = [f"{date_field} BETWEEN %s AND %s"]
        params = [filters["date_from"], filters["date_to"]]

        if filters.get("vendor_id"):
            conditions.append("l.vendor_id = %s")
            params.append(filters["vendor_id"])
        if filters.get("property_id"):
            conditions.append(f"{property_field} = %s")
            params.append(filters["property_id"])
        if include_account and filters.get("account_code"):
            conditions.append(f"{account_field} = %s")
            params.append(filters["account_code"])

        return " AND ".join(conditions), params

    def _run_report_query(self, title: str, query: str, params: Sequence[Any], filters: Dict[str, Any]):
        columns, rows = self._run_query(query, params)
        summary = self._compute_summary(columns, rows)
        return {
            "title": title,
            "columns": columns,
            "rows": rows,
            "summary": summary,
            "date_from": filters.get("date_from", ""),
            "date_to": filters.get("date_to", ""),
        }

    def _normalize_payload(self, payload, filters, report_name):
        columns = payload.get("columns") or []
        rows = payload.get("rows") or []
        summary = payload.get("summary") or self._compute_summary(columns, rows)
        return {
            "title": payload.get("title") or report_name,
            "columns": columns,
            "rows": rows,
            "summary": summary,
            "date_from": payload.get("date_from") or filters.get("date_from", ""),
            "date_to": payload.get("date_to") or filters.get("date_to", ""),
            "meta_lines": payload.get("meta_lines") or [],
        }

    def _generate_report(self, tab_key):
        filters = self._validate_filters(tab_key)
        if not filters:
            return

        state = self.tabs[tab_key]
        report_name = state["report_var"].get()

        if report_name == "كشف حساب تفصيلي":
            # Always return to dashboard before opening parameters dialog.
            self._back_to_reports_dashboard()
            self._open_account_statement_dialog()
            return

        handler = state["reports"].get(report_name)
        if not handler:
            messagebox.showwarning("تنبيه", "لم يتم العثور على نوع التقرير")
            return

        try:
            payload = handler(filters)
            _, has_rows = self._present_payload(tab_key, payload, filters, report_name)
        except ValueError as exc:
            messagebox.showwarning("تنبيه", str(exc))
            return
        except Exception as exc:
            messagebox.showerror("خطأ", f"تعذر إنشاء التقرير: {exc}")
            return

        if not has_rows:
            messagebox.showinfo("نتيجة", "لا توجد بيانات مطابقة للفلاتر المحددة")

    def _lookup_display_by_prefix(self, values, prefix):
        token = str(prefix or "").strip()
        if not token:
            return ""
        for value in values or []:
            if self._extract_prefix(value) == token:
                return value
        return token

    def _read_date_widget(self, widget):
        if hasattr(widget, "entry"):
            return widget.entry.get().strip()
        try:
            return widget.get().strip()
        except Exception:
            return ""

    def _build_date_widget(self, parent, initial_value):
        start_date = date.today()
        try:
            start_date = datetime.strptime(initial_value, "%Y-%m-%d").date()
        except Exception:
            pass
        if DateEntry is not None:
            try:
                widget = DateEntry(parent, dateformat="%Y-%m-%d", firstweekday=6, startdate=start_date, bootstyle="secondary", width=15)
                widget.entry.configure(justify="center")
                widget.entry.delete(0, tk.END)
                widget.entry.insert(0, initial_value)
                return widget
            except Exception:
                pass
        return ttk.Entry(parent, justify="center", font=self.base_font)

    def _table_columns(self, schema, table):
        key = f"{schema}.{table}"
        cached = self._lookup_cache.get(key)
        if cached is not None:
            return cached
        query = """
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_schema = %s AND table_name = %s
        """
        try:
            _, rows = self._run_query(query, [schema, table])
            columns = {row[0]: row[1] for row in rows}
        except Exception:
            columns = {}
        self._lookup_cache[key] = columns
        return columns

    def _resolve_posted_flag(self):
        cached = self._lookup_cache.get("posted_flag")
        if cached is not None:
            return cached
        candidates = (("v", "vouchers"), ("l", "ledger"))
        names = ("is_posted", "posted", "post_status", "is_posting")
        for alias, table in candidates:
            columns = self._table_columns("finance", table)
            for name in names:
                if columns.get(name) == "boolean":
                    self._lookup_cache["posted_flag"] = (alias, name)
                    return self._lookup_cache["posted_flag"]
        self._lookup_cache["posted_flag"] = None
        return None

    def _posted_clause(self, include_unposted):
        if include_unposted:
            return ""
        flag = self._resolve_posted_flag()
        if not flag:
            return ""
        alias, column = flag
        return f" AND COALESCE({alias}.{column}, TRUE) = TRUE"

    def _toggle_account_combo_state(self, state):
        combo = state["account_combo"]
        combo.configure(state="disabled" if state["all_accounts_var"].get() else "normal")

    # =====================================================
    # External API
    # =====================================================

    def open_report(self, report_name):
        route = self.report_to_route.get(report_name)
        if not route:
            self.notebook.select(0)
            return

        tab_key, option_name = route
        if tab_key == "dialog":
            # Dialog must open over the dashboard state only.
            self._back_to_reports_dashboard()
            self._open_account_statement_dialog()
            return

        self._back_to_reports_dashboard()

        tab_index = self.tab_order.index(tab_key)
        self.notebook.select(tab_index)

        state = self.tabs[tab_key]
        state["report_var"].set(option_name)
        self._generate_report(tab_key)


def open_report(master, report_name):
    """
    Factory helper used by the main app to open a report screen directly.
    """
    screen = ReportsScreen(master)
    if report_name:
        screen.open_report(report_name)
    return screen

